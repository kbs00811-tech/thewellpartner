"""
우리 결과 (result.xlsx) vs 완성 폼 (근퇴 테스트.xlsx) cell-by-cell 비교
"""
import sys
import io
import os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import load_workbook
from collections import defaultdict

OUR_PATH = "C:/tmp/lty_test/result.xlsx"
GOLDEN_PATH = "C:/Users/USER/OneDrive/Desktop/근퇴 완성/근퇴 테스트.xlsx"

# 우리 결과는 data_only=False (직원명은 수식 텍스트 — 비교 안 함, 행 번호로 매핑)
# 완성 폼은 data_only=True (직원명 평가값 가져옴)
wb_our = load_workbook(OUR_PATH, data_only=False, read_only=False)
wb_golden = load_workbook(GOLDEN_PATH, data_only=True, read_only=True)

ws_our = wb_our["근태 ( 3월 )"]
ws_golden = wb_golden["근태 ( 3월 )"]


def get_employees_from_evaluated(ws):
    """data_only=True ws에서 직원명 → 기본 행 매핑"""
    result = {}
    BASIC = {"기본", "기본근무", "정상"}
    for r in range(1, ws.max_row + 1):
        d = ws.cell(row=r, column=4).value
        f = ws.cell(row=r, column=6).value
        if d and isinstance(d, str) and isinstance(f, str) and f.strip() in BASIC:
            result[d.strip()] = r
    return result


# 완성 폼에서 직원명 → 행 (양쪽 파일이 같은 구조 = 같은 행)
golden_emps = get_employees_from_evaluated(ws_golden)
our_emps = golden_emps  # 같은 행 번호 사용

print(f"우리 결과 직원: {len(our_emps)}명")
print(f"완성 폼 직원: {len(golden_emps)}명")
print()

# 6행 라벨
ROW_LABELS = ["기본", "연장", "심야", "특근", "특잔", "지각조퇴"]

# 모든 cell 비교
diff_count = 0
match_count = 0
diffs_by_employee = defaultdict(list)

# 모든 직원 (양쪽 union)
all_names = sorted(set(our_emps.keys()) | set(golden_emps.keys()))

for name in all_names:
    our_r = our_emps.get(name)
    golden_r = golden_emps.get(name)

    if our_r is None:
        diffs_by_employee[name].append({"type": "missing_in_ours"})
        diff_count += 1
        continue
    if golden_r is None:
        diffs_by_employee[name].append({"type": "missing_in_golden"})
        diff_count += 1
        continue

    # 6행 × 31일 + AM (월말 보정)
    for offset, label in enumerate(ROW_LABELS):
        our_row = our_r + offset
        golden_row = golden_r + offset

        for day in range(1, 32):
            col = 7 + day  # H=8(1일), AL=38(31일)
            our_v = ws_our.cell(row=our_row, column=col).value
            golden_v = ws_golden.cell(row=golden_row, column=col).value

            # 정규화: None == "" == 0 (간주)
            def norm(v):
                if v in (None, ""):
                    return None
                if isinstance(v, float) and v == int(v):
                    return int(v)
                return v

            ov = norm(our_v)
            gv = norm(golden_v)
            if ov != gv:
                diffs_by_employee[name].append({
                    "type": "cell_diff",
                    "label": label,
                    "day": day,
                    "col": col,
                    "our": ov,
                    "golden": gv,
                })
                diff_count += 1
            else:
                match_count += 1

        # AM열 (월말 보정)
        col_am = 39
        our_v = ws_our.cell(row=our_row, column=col_am).value
        golden_v = ws_golden.cell(row=golden_row, column=col_am).value
        def norm(v):
            if v in (None, ""):
                return None
            if isinstance(v, float) and v == int(v):
                return int(v)
            return v
        ov = norm(our_v)
        gv = norm(golden_v)
        if ov != gv:
            diffs_by_employee[name].append({
                "type": "cell_diff",
                "label": f"{label}_AM",
                "day": "AM",
                "col": col_am,
                "our": ov,
                "golden": gv,
            })
            diff_count += 1
        else:
            match_count += 1

# 요약 출력
print(f"=" * 70)
print(f"비교 결과")
print(f"=" * 70)
print(f"  일치: {match_count}")
print(f"  불일치: {diff_count}")
print()

# 직원별 차이점 출력 (불일치 많은 순)
print(f"=" * 70)
print(f"직원별 차이점 ({len(diffs_by_employee)}명 영향)")
print(f"=" * 70)
sorted_emps = sorted(diffs_by_employee.items(), key=lambda x: -len(x[1]))
for name, diffs in sorted_emps[:30]:
    print(f"\n--- {name} ({len(diffs)}건) ---")
    for d in diffs[:20]:  # 직원당 최대 20건
        if d["type"] == "cell_diff":
            print(f"  {d['label']:6} day={d['day']:>2} (col={d['col']}): our={repr(d['our']):12} golden={repr(d['golden'])}")
        else:
            print(f"  {d['type']}")
    if len(diffs) > 20:
        print(f"  ... +{len(diffs)-20} more")

# 차이점 유형 분류
print(f"\n{'=' * 70}")
print(f"차이점 유형별 분류")
print(f"=" * 70)
type_count = defaultdict(int)
pattern_count = defaultdict(int)
for diffs in diffs_by_employee.values():
    for d in diffs:
        type_count[d.get("type", "?")] += 1
        if d.get("type") == "cell_diff":
            our_v = d["our"]
            golden_v = d["golden"]
            pattern = f"our={type(our_v).__name__} → golden={type(golden_v).__name__}"
            if our_v is None and golden_v is not None:
                pattern = f"NONE → {repr(golden_v)[:20]}"
            elif our_v is not None and golden_v is None:
                pattern = f"{repr(our_v)[:20]} → NONE"
            pattern_count[pattern] += 1

for t, c in sorted(type_count.items(), key=lambda x: -x[1]):
    print(f"  {t}: {c}")
print()
for p, c in sorted(pattern_count.items(), key=lambda x: -x[1])[:20]:
    print(f"  {c:5}건  {p}")
