"""
급여명세서 발송 API
- 결과 xlsx 업로드 → 직원 목록 + 명세서 데이터 추출
- 직원별 명세서 토큰 URL 발급
- 솔라피(SOLAPI) API로 SMS / 알림톡 발송
"""
from __future__ import annotations
import os
import json
import secrets
import tempfile
import time
import traceback
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, List

from fastapi import APIRouter, UploadFile, File, Form, HTTPException, Header
from pydantic import BaseModel
from openpyxl import load_workbook

from ..auth import verify_admin_token

router = APIRouter(prefix="/api/payslip", tags=["payslip"])


# === 토큰 저장소 — 메모리 + 디스크 백업 (24시간 유효) ===
# Render 컨테이너 재시작/재배포 시 메모리 토큰 소실 방지
# Render Standard는 idle 슬립 없으나 배포 시 재시작 → 디스크 백업으로 복원
TOKEN_STORE: dict = {}
TOKEN_TTL = 86400  # 24시간

# Render 영구 디스크 권장 경로. 미설정 시 임시 디스크(/tmp)
TOKEN_STORE_PATH = os.environ.get("PAYSLIP_TOKEN_STORE", "/tmp/payslip_tokens.pkl")


def _log(msg: str):
    import sys
    print(f"[PAYSLIP] {msg}", flush=True)
    sys.stdout.flush()


def _save_tokens_to_disk():
    """TOKEN_STORE 디스크 백업 (배포/재시작 후 복원용)"""
    try:
        import pickle
        # bytes 데이터 포함이라 pickle 사용 (JSON 불가)
        with open(TOKEN_STORE_PATH, "wb") as f:
            pickle.dump(TOKEN_STORE, f, protocol=pickle.HIGHEST_PROTOCOL)
    except Exception as e:
        _log(f"token save failed (non-fatal): {type(e).__name__}: {e}")


def _load_tokens_from_disk():
    """시작 시 디스크에서 토큰 복원"""
    try:
        if not os.path.exists(TOKEN_STORE_PATH):
            return
        import pickle
        with open(TOKEN_STORE_PATH, "rb") as f:
            data = pickle.load(f)
        # 만료된 것은 자동 제외
        now = time.time()
        loaded = 0
        for k, v in data.items():
            if v.get("expires_at", 0) > now:
                TOKEN_STORE[k] = v
                loaded += 1
        if loaded:
            _log(f"token store restored from disk: {loaded} active tokens")
    except Exception as e:
        _log(f"token load failed (non-fatal): {type(e).__name__}: {e}")


# 모듈 import 시 디스크에서 자동 복원
_load_tokens_from_disk()


def _cleanup_expired_tokens():
    """만료된 토큰 제거 (메모리 관리)"""
    now = time.time()
    expired = [k for k, v in TOKEN_STORE.items() if v.get("expires_at", 0) < now]
    if expired:
        for k in expired:
            TOKEN_STORE.pop(k, None)
        _save_tokens_to_disk()


class EmployeeInfo(BaseModel):
    no: int
    name: str
    phone: Optional[str] = None
    hire_date: Optional[str] = None
    resign_date: Optional[str] = None
    hourly_rate: Optional[float] = None
    bank: Optional[str] = None
    account: Optional[str] = None


def _extract_employees_from_xlsx(xlsx_path: str) -> List[EmployeeInfo]:
    """xlsx의 DB 시트에서 직원 명단 추출 (이름, 연락처, 시급, 은행/계좌)"""
    wb = load_workbook(xlsx_path, data_only=True, read_only=True, keep_links=False)
    employees: List[EmployeeInfo] = []
    try:
        if "DB" not in wb.sheetnames:
            return employees
        ws = wb["DB"]
        # R1=헤더, R2~ 직원 데이터
        for r in range(2, ws.max_row + 1):
            no_val = ws.cell(r, 1).value
            name_val = ws.cell(r, 2).value
            if not name_val or not isinstance(name_val, str):
                continue
            phone = ws.cell(r, 7).value
            hire = ws.cell(r, 5).value
            resign = ws.cell(r, 6).value
            rate = ws.cell(r, 3).value
            bank = ws.cell(r, 11).value
            account = ws.cell(r, 12).value
            try:
                no_int = int(no_val) if no_val else r - 1
            except (ValueError, TypeError):
                no_int = r - 1
            employees.append(EmployeeInfo(
                no=no_int,
                name=name_val.strip(),
                phone=str(phone).strip() if phone else None,
                hire_date=hire.strftime("%Y-%m-%d") if isinstance(hire, datetime) else (str(hire) if hire else None),
                resign_date=resign.strftime("%Y-%m-%d") if isinstance(resign, datetime) else (str(resign) if resign else None),
                hourly_rate=float(rate) if rate else None,
                bank=str(bank).strip() if bank else None,
                account=str(account).strip() if account else None,
            ))
    finally:
        wb.close()
    return employees


@router.post("/employees")
async def list_employees(
    excel: UploadFile = File(...),
    x_admin_token: Optional[str] = Header(None, alias="X-Admin-Token"),
    authorization: Optional[str] = Header(None),
):
    """결과 xlsx 업로드 → 직원 명단 + 연락처 반환 (어드민이 발송 대상 선택용)"""
    verify_admin_token(x_admin_token, authorization)
    tmp_dir = Path(tempfile.mkdtemp(prefix="payslip_"))
    try:
        xlsx_path = tmp_dir / (excel.filename or "in.xlsx")
        with open(xlsx_path, "wb") as f:
            f.write(await excel.read())
        emps = _extract_employees_from_xlsx(str(xlsx_path))
        _log(f"list_employees: {len(emps)}명")
        return {"employees": [e.model_dump() for e in emps]}
    finally:
        import shutil
        shutil.rmtree(tmp_dir, ignore_errors=True)


class SendRequest(BaseModel):
    company_id: Optional[str] = "lty"
    year: int
    month: int
    channel: str = "sms"  # 'sms' | 'lms' | 'alimtalk'
    employee_nos: List[int]  # 발송 대상 직원 번호 목록


@router.post("/send")
async def send_payslip(
    excel: UploadFile = File(...),
    payload: str = Form(...),  # JSON 직렬화된 SendRequest
    x_admin_token: Optional[str] = Header(None, alias="X-Admin-Token"),
    authorization: Optional[str] = Header(None),
):
    """
    명세서 발송 (Phase D — 솔라피 연동).
    현재는 토큰 URL만 발급. 실제 솔라피 발송은 API 키 환경변수 등록 후 활성화.
    """
    verify_admin_token(x_admin_token, authorization)
    _cleanup_expired_tokens()

    try:
        req = SendRequest(**json.loads(payload))
    except Exception as e:
        raise HTTPException(400, f"payload 파싱 실패: {e}")

    api_key = os.environ.get("SOLAPI_API_KEY", "").strip()
    api_secret = os.environ.get("SOLAPI_API_SECRET", "").strip()
    sender = os.environ.get("SOLAPI_SENDER", "").strip()
    solapi_ready = bool(api_key and api_secret and sender)

    # 업로드 파일 저장
    tmp_dir = Path(tempfile.mkdtemp(prefix="payslip_send_"))
    try:
        xlsx_path = tmp_dir / (excel.filename or "in.xlsx")
        with open(xlsx_path, "wb") as f:
            f.write(await excel.read())

        # 전체 직원 정보 추출
        all_emps = _extract_employees_from_xlsx(str(xlsx_path))
        emps_by_no = {e.no: e for e in all_emps}
        targets = [emps_by_no[n] for n in req.employee_nos if n in emps_by_no]
        _log(f"send: 대상 {len(targets)}명 / channel={req.channel}")

        # xlsx 한 번만 읽어 메모리에 보관 (직원당 반복 X)
        with open(xlsx_path, "rb") as f:
            xlsx_bytes_cached = f.read()

        # 직원별 토큰 발급
        results = []
        base_url = os.environ.get("PAYSLIP_BASE_URL", "https://thewellpartner.com").rstrip("/")
        for emp in targets:
            token = secrets.token_urlsafe(16)
            TOKEN_STORE[token] = {
                "employee_no": emp.no,
                "employee_name": emp.name,
                "year": req.year,
                "month": req.month,
                "company_id": req.company_id,
                "xlsx_bytes": xlsx_bytes_cached,  # 캐시된 bytes 공유
                "expires_at": time.time() + TOKEN_TTL,
            }
            url = f"{base_url}/payslip/{token}"

            # 솔라피 발송 (API 키 있을 때만)
            send_status = "skipped_no_api_key"
            error = None
            if not solapi_ready:
                if not api_key:
                    error = "SOLAPI_API_KEY env not set"
                elif not api_secret:
                    error = "SOLAPI_API_SECRET env not set"
                elif not sender:
                    error = "SOLAPI_SENDER env not set"
            elif not emp.phone:
                send_status = "no_phone"
                error = "직원 연락처 없음"
            else:
                try:
                    send_status = _send_solapi(
                        api_key, api_secret, sender,
                        to_phone=emp.phone,
                        channel=req.channel,
                        text=_build_message(emp.name, req.year, req.month, url),
                    )
                except Exception as e:
                    error = f"{type(e).__name__}: {str(e)[:300]}"
                    send_status = "failed"
                    _log(f"  send fail {emp.name}: {error}")

            results.append({
                "no": emp.no,
                "name": emp.name,
                "phone": emp.phone,
                "url": url,
                "status": send_status,
                "error": error,
            })

        # 모든 토큰 추가 + 발송 완료 후 디스크 백업 (배포/재시작 후 복원용)
        _save_tokens_to_disk()

        return {
            "ok": True,
            "solapi_ready": solapi_ready,
            "channel": req.channel,
            "total": len(targets),
            "results": results,
        }
    except HTTPException:
        raise
    except Exception as e:
        _log(f"send ERROR: {type(e).__name__}: {e}")
        _log(traceback.format_exc())
        raise HTTPException(500, f"발송 실패: {type(e).__name__}: {e}")
    finally:
        import shutil
        shutil.rmtree(tmp_dir, ignore_errors=True)


def _build_message(name: str, year: int, month: int, url: str) -> str:
    """SMS / 알림톡 본문 — 솔라피 알림톡 템플릿과 동일하게"""
    return (
        f"[더웰파트너 급여명세서]\n"
        f"{name} 님, {year}년 {month}월분 급여명세서가 도착했습니다.\n"
        f"명세서 보기: {url}\n"
        f"(24시간 유효)"
    )


def _send_solapi(
    api_key: str,
    api_secret: str,
    sender: str,
    to_phone: str,
    channel: str,
    text: str,
) -> str:
    """솔라피 SDK 5.x로 발송. 성공 시 'sent' 반환.

    공식 SDK는 type 파라미터를 받지 않음 — 텍스트 길이로 자동 SMS/LMS 결정.
    알림톡은 KakaoOption (pf_id + template_id) 필요 — 환경변수 등록 안 되면 SMS로 폴백.
    """
    try:
        from solapi import SolapiMessageService
        from solapi.model import RequestMessage
    except ImportError:
        return "sdk_not_installed"

    msg_service = SolapiMessageService(api_key=api_key, api_secret=api_secret)
    sender_clean = sender.replace("-", "").replace(" ", "")
    to_clean = to_phone.replace("-", "").replace(" ", "")

    kakao_option = None
    if channel == "alimtalk":
        pf_id = os.environ.get("SOLAPI_PF_ID", "").strip()
        template_id = os.environ.get("SOLAPI_TEMPLATE_ID", "").strip()
        if pf_id and template_id:
            try:
                from solapi.model.kakao.kakao_option import KakaoOption
                kakao_option = KakaoOption(pf_id=pf_id, template_id=template_id)
            except ImportError:
                kakao_option = None
        # 알림톡 옵션 없으면 SMS로 폴백
        # (사용자가 KakaoBiz 채널/템플릿 등록 안 한 단계)

    # RequestMessage 생성 — type 인자 사용 안 함 (텍스트 길이로 자동)
    msg_kwargs = {
        "from_": sender_clean,
        "to": to_clean,
        "text": text,
    }
    if kakao_option is not None:
        msg_kwargs["kakao_options"] = kakao_option

    message = RequestMessage(**msg_kwargs)
    response = msg_service.send(message)
    # 응답 검증 — registered_failed > 0 이면 실패
    try:
        failed = getattr(response.group_info.count, "registered_failed", 0) or 0
        if failed > 0:
            raise RuntimeError(f"solapi registered_failed={failed}")
    except AttributeError:
        pass  # 응답 구조 다른 경우 그냥 통과
    return "sent"


@router.get("/data/{token}")
async def get_payslip_data(token: str):
    """직원용 명세서 페이지가 토큰으로 데이터 조회 (인증 불필요, 토큰 자체가 인증)"""
    _cleanup_expired_tokens()
    info = TOKEN_STORE.get(token)
    if not info:
        raise HTTPException(404, "유효하지 않거나 만료된 명세서 링크")

    if info["expires_at"] < time.time():
        TOKEN_STORE.pop(token, None)
        raise HTTPException(410, "명세서 링크가 만료되었습니다 (24시간 유효)")

    # 직원 데이터 추출
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    try:
        tmp.write(info["xlsx_bytes"])
        tmp.close()
        data = _extract_payslip_data(tmp.name, info["employee_no"], info["year"], info["month"])
        data["employee_no"] = info["employee_no"]
        data["employee_name"] = info["employee_name"]
        return data
    finally:
        try:
            os.unlink(tmp.name)
        except OSError:
            pass


def _extract_payslip_data(xlsx_path: str, employee_no: int, year: int, month: int) -> dict:
    """양식 xlsx에서 특정 직원의 명세서 데이터 추출.

    지급 내역 시트 컬럼 매핑 (D=직원명 키, R10=헤더, R11+=데이터):
      P=기본급(16) Q=잔업(17) R=심야(18) S=특근(19) T=특잔(20) U=지각조퇴(21)
      V=식대(22) W=교통비(23) X=연차(24) Y=원단(25) AA=퇴직금(27)
      AC=직접비소계(29)
      AD=국민연금(30) AE=건강보험(31) AF=노인장기(32) AG=고용보험(33) AH=4대보험합계(34)
      AJ=소득세(36) AK=가불금(37) AL=실수령(38)
    """
    def _num(v):
        """숫자 정규화 (수식 캐시 빈 결과는 0)"""
        if v is None or v == "":
            return 0
        try:
            return float(v)
        except (ValueError, TypeError):
            return 0

    def _str(v):
        if v is None:
            return ""
        if isinstance(v, datetime):
            return v.strftime("%Y-%m-%d")
        return str(v).strip()

    wb = load_workbook(xlsx_path, data_only=True, read_only=False, keep_links=False)
    try:
        result = {
            "year": year,
            "month": month,
            "soc": {},
            "pay": {},
            "deduct": {},
            "attendance": [],  # [{day, basic, ot, night, sat, sat_ot, late}]
        }

        # === 1. DB 시트에서 직원 기본 정보 ===
        emp_name = ""
        if "DB" in wb.sheetnames:
            ws = wb["DB"]
            row = employee_no + 1  # NO=1 → R2
            if 2 <= row <= ws.max_row:
                emp_name = _str(ws.cell(row, 2).value)
                result["soc"] = {
                    "name": emp_name,
                    "hire_date": _str(ws.cell(row, 5).value),
                    "resign_date": _str(ws.cell(row, 6).value),
                    "rate": _num(ws.cell(row, 3).value),
                    "bank": _str(ws.cell(row, 11).value),
                    "account": _str(ws.cell(row, 12).value),
                }

        # === 2. 지급 내역 시트에서 직원 행 찾고 급여/공제 추출 ===
        # 우선 캐시된 값 시도 (Excel 한번 열어 저장한 경우)
        pay_sheet_name = next((s for s in wb.sheetnames if "지급" in s), None)
        pay_values_from_cache = False
        if pay_sheet_name and emp_name:
            ws_pay = wb[pay_sheet_name]
            target_row = None
            # 두 가지 방식 시도: 1) D열 값 매칭, 2) NO 기반 위치 (R10+no)
            try:
                for r in range(11, min(ws_pay.max_row + 1, 50)):
                    d_val = _str(ws_pay.cell(r, 4).value)
                    if d_val == emp_name:
                        target_row = r
                        break
            except Exception:
                pass
            if target_row is None:
                # NO 기반 fallback: R11=NO 1
                target_row = 10 + employee_no

            if target_row and target_row <= ws_pay.max_row:
                r = target_row
                cached_basic = _num(ws_pay.cell(r, 16).value)
                if cached_basic > 0:
                    pay_values_from_cache = True
                    result["pay"] = {
                        "basic": cached_basic,
                        "overtime": _num(ws_pay.cell(r, 17).value),
                        "night": _num(ws_pay.cell(r, 18).value),
                        "saturday": _num(ws_pay.cell(r, 19).value),
                        "sat_ot": _num(ws_pay.cell(r, 20).value),
                        "late": _num(ws_pay.cell(r, 21).value),
                        "meal": _num(ws_pay.cell(r, 22).value),
                        "transport": _num(ws_pay.cell(r, 23).value),
                        "annual_leave": _num(ws_pay.cell(r, 24).value),
                        "shift": _num(ws_pay.cell(r, 25).value),
                        "retirement": _num(ws_pay.cell(r, 27).value),
                        "subtotal": _num(ws_pay.cell(r, 29).value),
                    }
                    result["deduct"] = {
                        "pension": _num(ws_pay.cell(r, 30).value),
                        "health": _num(ws_pay.cell(r, 31).value),
                        "longterm": _num(ws_pay.cell(r, 32).value),
                        "employment": _num(ws_pay.cell(r, 33).value),
                        "insurance_total": _num(ws_pay.cell(r, 34).value),
                        "income_tax": _num(ws_pay.cell(r, 36).value),
                        "loan": _num(ws_pay.cell(r, 37).value),
                    }
                    result["net_pay"] = _num(ws_pay.cell(r, 38).value)
                    result["deduct_total"] = (
                        result["deduct"]["insurance_total"]
                        + result["deduct"]["income_tax"]
                        + result["deduct"]["loan"]
                    )

        # === 3. 근태 시트에서 일자별 출퇴근 + 시간 합계 (캐시 없을 때 직접 계산용) ===
        att_sheet = next((s for s in wb.sheetnames if f"근태 ( {month}" in s), None)
        if not att_sheet:
            att_sheet = next((s for s in wb.sheetnames if "근태" in s and "월" in s), None)

        hours = {"basic": 0.0, "overtime": 0.0, "night": 0.0, "saturday": 0.0, "sat_ot": 0.0, "late": 0.0}
        if att_sheet and emp_name:
            ws_att = wb[att_sheet]
            BASIC = {"기본", "기본근무", "정상"}
            target_base_row = None
            # NO 기반 위치 추정 — 양식: R7=NO 1 기본행, 6행 블록
            estimated_row = 1 + (employee_no - 1) * 6 + 6  # NO=1 → R7
            # 우선 NO 기반, 안 맞으면 D열 매칭
            try:
                f_val = _str(ws_att.cell(estimated_row, 6).value)
                if f_val in BASIC:
                    target_base_row = estimated_row
            except Exception:
                pass
            if target_base_row is None:
                for r in range(1, min(ws_att.max_row + 1, 250)):
                    d_val = _str(ws_att.cell(r, 4).value)
                    f_val = _str(ws_att.cell(r, 6).value)
                    if d_val == emp_name and f_val in BASIC:
                        target_base_row = r
                        break

            if target_base_row:
                LABELS = ["basic", "overtime", "night", "saturday", "sat_ot", "late"]
                for day in range(1, 32):
                    col = 7 + day  # H=8(1일), AL=38(31일)
                    day_data = {"day": day}
                    has_any = False
                    for off, key in enumerate(LABELS):
                        v = ws_att.cell(target_base_row + off, col).value
                        if v not in (None, ""):
                            day_data[key] = v
                            has_any = True
                            # 시간 합계 (숫자만)
                            try:
                                hours[key] += float(v)
                            except (ValueError, TypeError):
                                pass
                    if has_any:
                        result["attendance"].append(day_data)

        result["hours"] = hours

        # === 4. 캐시 값 없으면 직접 계산 (시급 × 시간 × 배율) ===
        if not pay_values_from_cache and result["soc"].get("rate"):
            rate = result["soc"]["rate"]
            def rounddown(v, digits=-1):
                """엑셀 ROUNDDOWN(-1) = 10원 단위 내림"""
                if digits == -1:
                    return int(v // 10) * 10
                return int(v)

            basic_pay = rounddown(rate * hours["basic"])
            ot_pay = rounddown(rate * hours["overtime"] * 1.5)
            night_pay = rounddown(rate * hours["night"] * 0.5)
            sat_pay = rounddown(rate * hours["saturday"] * 1.5)
            sat_ot_pay = rounddown(rate * hours["sat_ot"] * 2)
            late_pay = rounddown(rate * hours["late"])  # 지각조퇴는 음수
            subtotal = basic_pay + ot_pay + night_pay + sat_pay + sat_ot_pay + late_pay

            result["pay"] = {
                "basic": basic_pay,
                "overtime": ot_pay,
                "night": night_pay,
                "saturday": sat_pay,
                "sat_ot": sat_ot_pay,
                "late": late_pay,
                "subtotal": subtotal,
                # 식대/교통비/연차/원단/퇴직금은 캐시 없으면 0 (회사별 룰)
                "meal": 0, "transport": 0, "annual_leave": 0,
                "shift": 0, "retirement": 0,
            }

            # 4대 보험 자체 계산 (대략)
            pension = rounddown(subtotal * 0.0475 if subtotal > 0 else 0)
            health = rounddown(subtotal * 0.03595 if subtotal > 0 else 0)
            longterm = rounddown(health * 0.1314 if health > 0 else 0)
            employment = rounddown(subtotal * 0.009 if subtotal > 0 else 0)
            insurance_total = pension + health + longterm + employment

            result["deduct"] = {
                "pension": pension,
                "health": health,
                "longterm": longterm,
                "employment": employment,
                "insurance_total": insurance_total,
                "income_tax": 0,
                "loan": 0,
            }
            result["deduct_total"] = insurance_total
            result["net_pay"] = subtotal - insurance_total
            result["data_source"] = "계산값 (Excel 미평가)"
        else:
            result["data_source"] = "양식 평가값"

        return result
    finally:
        wb.close()
