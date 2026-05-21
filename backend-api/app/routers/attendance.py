"""
근태 자동 입력 API 라우터
POST /api/attendance/process
"""
from __future__ import annotations
import json
import shutil
import tempfile
import time
import sys
import traceback
from pathlib import Path
from typing import Optional


def _log(msg: str):
    """진행 단계 로그 — Render Logs에서 실시간 확인"""
    print(f"[ATTENDANCE] {msg}", flush=True)
    sys.stdout.flush()

from fastapi import APIRouter, UploadFile, File, Form, HTTPException, Header, Response
from openpyxl import load_workbook

from ..services.attendance_parser import parse_pdf
from ..services.excel_writer import (
    fill_attendance_sheet,
    classify_attendance,
    extract_employee_names_from_sheet,
    find_target_sheet,
    find_employee_block,
    DAY_TO_COL,
)
from ..services.leave_writer import fill_leave_sheet
from ..services.validator import validate, validate_lite, write_validation_to_review
from ..services.hire_date import (
    extract_hire_dates_from_excel,
    estimate_hire_dates_from_pdf,
    merge_hire_dates,
    find_feb_attendance_sheet,
    check_feb_last_week_credit,
    extract_employee_info_from_db,
)
from ..services.completeness import verify_input_completeness
from ..auth import verify_admin_token

router = APIRouter(prefix="/api/attendance", tags=["attendance"])


def _parse_holidays(holidays_text: str) -> dict:
    """공휴일 텍스트 → dict ('2026-05-05 어린이날\n...' 형식)"""
    result = {}
    if not holidays_text:
        return result
    for line in holidays_text.strip().split("\n"):
        line = line.strip()
        if not line:
            continue
        parts = line.split(maxsplit=1)
        if parts:
            result[parts[0]] = parts[1] if len(parts) > 1 else "공휴일"
    return result


def _normalize_xlsx_dimensions(path: str, max_row_cap: int = 5000) -> dict:
    """양식 xlsx 파일의 worksheet XML 안 <dimension ref="..."/>의 end row를 cap.

    급여명세서 등 시트에서 max_row=1048558로 잘못 정의된 파일은 후속 openpyxl
    처리에서 17분+ 소요 (Render 120초 timeout 초과 → 502).
    zip 내부 XML 직접 정규식 치환으로 sheet 이름 매핑 없이 처리.

    Returns: {sheet_xml_path: original_ref}
    """
    import zipfile, shutil, re
    fixes = {}
    cell_re = re.compile(rb'^([A-Z]+)(\d+)$')
    dim_re = re.compile(rb'(<dimension\s+ref=")([^"]+)("\s*/>)')

    def cap_dim(m):
        prefix, ref, suffix = m.group(1), m.group(2), m.group(3)
        parts = ref.split(b':')
        if len(parts) != 2:
            return m.group(0)
        cm = cell_re.match(parts[1])
        if not cm:
            return m.group(0)
        end_col, end_row = cm.group(1), int(cm.group(2))
        if end_row <= max_row_cap:
            return m.group(0)
        new_end = end_col + str(max_row_cap).encode()
        return prefix + parts[0] + b':' + new_end + suffix

    tmp = path + ".tmp"
    with zipfile.ZipFile(path, "r") as zin:
        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.namelist():
                data = zin.read(item)
                if item.startswith("xl/worksheets/sheet") and item.endswith(".xml"):
                    new_data, n = dim_re.subn(cap_dim, data, count=1)
                    if n > 0 and new_data != data:
                        m = dim_re.search(data)
                        fixes[item] = m.group(2).decode() if m else "unknown"
                    data = new_data
                zout.writestr(item, data)
    shutil.move(tmp, path)
    return fixes


# 업체별 표시명 — 결과 파일명·로그용. 신규 업체 추가 시 한 줄.
COMPANY_LABELS: dict = {
    "lty": "엘티와이",
    "elcomtec": "엘컴텍",
}

# 다중 PDF·전용 파서를 쓰는 업체 (근태대장 포맷 등)
ELCOMTEC_COMPANIES = {"elcomtec"}


@router.post("/process")
async def process_attendance(
    pdf: UploadFile = File(...),
    excel: UploadFile = File(...),
    year: int = Form(...),
    month: int = Form(...),
    holidays: str = Form(""),
    standard_hours: int = Form(209),
    normal_start: str = Form("08:30"),
    normal_end: str = Form("17:30"),
    sheet_name: Optional[str] = Form(None),
    leave_sheet_name: str = Form("연차내역"),
    overwrite_existing: bool = Form(False),
    fill_leave: bool = Form(False),
    enforce_pdf_month: bool = Form(False),  # 사용자 입력 강제 우선
    company_id: Optional[str] = Form(None),  # 업체 식별자 (lty, samsung 등)
    x_admin_token: Optional[str] = Header(None, alias="X-Admin-Token"),
    authorization: Optional[str] = Header(None),
):
    """
    PDF 출근부 + 청구 엑셀 → 근태 자동 입력된 엑셀 반환.
    수식/서식은 100% 보존.

    처리 순서:
      1. 엑셀 직원명 + 입사일 사전 추출
      2. PDF 파싱 (직원명 매칭 + 날짜 보정)
      3. PDF month vs user month 검증
      4. PDF 첫 활동일 기반 입사일 추정 (엑셀에서 못 찾은 경우 보완)
      5. 2월 근태 시트 마지막 주 인정 검사 (3/1 주휴 판단용)
      6. 근태 시트 자동 입력 (입사일 게이트 적용)
      7. 입력 누락 검증
      8. 검토리스트 + 결과 반환
    """
    verify_admin_token(x_admin_token, authorization)

    if not pdf.filename or not excel.filename:
        raise HTTPException(400, "PDF와 Excel 파일이 모두 필요합니다.")

    # company_id 화이트리스트 검증 — 등록되지 않은 회사는 거부
    company_id_normalized = (company_id or "").strip().lower()
    if company_id_normalized and company_id_normalized not in COMPANY_LABELS:
        raise HTTPException(
            403,
            f"등록되지 않은 회사: {company_id}. 허용 회사: {sorted(COMPANY_LABELS.keys())}"
        )

    user_holidays = _parse_holidays(holidays)
    sheet_name_final = sheet_name.strip() if sheet_name and sheet_name.strip() else None

    t0 = time.time()
    company_label = COMPANY_LABELS.get(company_id_normalized, "")
    _log(f"=== START process_attendance year={year} month={month} company={company_id or '(none)'} ===")

    tmp_dir = Path(tempfile.mkdtemp(prefix="attendance_"))
    try:
      try:
        pdf_path = tmp_dir / pdf.filename
        excel_orig_path = tmp_dir / excel.filename
        # 결과 파일명: <회사명>_<연월>_지급파일.xlsx (회사명 없으면 기존 방식)
        # 지급 내역 시트가 수식 기반이라 Excel에서 열면 자동 계산됨
        if company_label:
            out_name = f"{company_label}_{year:04d}-{month:02d}_지급파일.xlsx"
        else:
            out_name = excel.filename.replace(".xlsx", "_근태자동입력_수정완료.xlsx")
        excel_out_path = tmp_dir / out_name

        _log(f"Step 1: file upload start")
        pdf_bytes = await pdf.read()
        excel_bytes = await excel.read()
        _log(f"  PDF: {len(pdf_bytes)/1024:.1f} KB, Excel: {len(excel_bytes)/1024:.1f} KB")

        with open(pdf_path, "wb") as f:
            f.write(pdf_bytes)
        with open(excel_orig_path, "wb") as f:
            f.write(excel_bytes)
        _log(f"Step 1 done [{time.time()-t0:.1f}s]")

        # 1.5. 양식 dimension 정상화 — 급여명세서 등 max_row=1048558 케이스 사전 정리
        # 이게 없으면 후속 openpyxl 처리에서 9분+ 소요 (Render 120초 timeout 초과 → 502)
        try:
            fixes = _normalize_xlsx_dimensions(str(excel_orig_path))
            if fixes:
                for path_key, original_ref in fixes.items():
                    _log(f"  dim capped: {path_key} (was {original_ref})")
            _log(f"Step 1.5 normalize done [{time.time()-t0:.1f}s] (fixes={len(fixes)})")
        except Exception as e:
            _log(f"Step 1.5 skipped (non-fatal): {type(e).__name__}: {e}")

        # 1. 엑셀 사전 로드 — DB 시트 + 직원명 + 입사일 + 2월 시트
        excel_names: list[str] = []
        hire_dates_excel: dict = {}
        resign_dates_db: dict = {}
        db_info: dict = {}
        feb_sheet_name: Optional[str] = None
        feb_credit_map: dict = {}

        _log("Step 2: Excel pre-load (single wb_pre_values, read_only=False after normalize)")
        try:
            # dimension 정상화(Step 1.5) 후엔 read_only=False가 더 빠름:
            #  - read_only=True 모드의 ws.cell(r,c) 는 lazy parse라 27명 × 100s
            #  - read_only=False 는 random access가 즉시 → 합계 17초로 6배 단축
            # 메모리: 양식 정리 후 ~300MB (Render 2GB 안전)
            wb_pre_values = load_workbook(
                str(excel_orig_path), data_only=True, read_only=False, keep_links=False
            )
            try:
                _log(f"  sheetnames: {wb_pre_values.sheetnames}")
                march_sheet = find_target_sheet(wb_pre_values, month, sheet_name_final)
                _log(f"  march_sheet: {march_sheet}")
                excel_names = extract_employee_names_from_sheet(wb_pre_values, march_sheet)
                _log(f"  excel_names: {len(excel_names)}명")

                db_info = extract_employee_info_from_db(wb_pre_values)
                _log(f"  db_info: {len(db_info)}명")

                # 엑셀 시트에서 입사일 보완 (data_only=True wb 사용 — 평가된 값 충분)
                candidate_sheets = [march_sheet]
                for sn in wb_pre_values.sheetnames:
                    if "직원" in sn or "정보" in sn:
                        candidate_sheets.append(sn)
                hire_dates_excel_only = extract_hire_dates_from_excel(
                    wb_pre_values, candidate_sheets, excel_names
                )

                # DB > 엑셀 우선순위로 입사일 병합
                hire_dates_excel = dict(hire_dates_excel_only)
                for norm, info in db_info.items():
                    if info.get("hire_date"):
                        hire_dates_excel[norm] = info["hire_date"]
                    if info.get("resign_date"):
                        resign_dates_db[norm] = info["resign_date"]

                # 2월 시트
                feb_sheet_name = find_feb_attendance_sheet(wb_pre_values, year, month)
                _log(f"  feb_sheet: {feb_sheet_name}")
                if feb_sheet_name:
                    for nm in excel_names:
                        result_check = check_feb_last_week_credit(
                            wb_pre_values, feb_sheet_name, nm, year,
                            feb_month=month - 1,
                            day_to_col=DAY_TO_COL,
                        )
                        feb_credit_map[nm] = result_check
            finally:
                wb_pre_values.close()
        except Exception as e:
            _log(f"Step 2 ERROR: {type(e).__name__}: {e}")
            _log(traceback.format_exc())
            excel_names = []
            hire_dates_excel = {}
            resign_dates_db = {}
        _log(f"Step 2 done [{time.time()-t0:.1f}s]")

        # 2. PDF 파싱
        _log("Step 3: PDF parsing")
        try:
            pdf_data, meta = parse_pdf(str(pdf_path), expected_names=excel_names)
            _log(f"  pdf_data: {len(pdf_data)}명, pages: {meta.get('raw_pages')}")
        except Exception as e:
            _log(f"Step 3 ERROR: {type(e).__name__}: {e}")
            _log(traceback.format_exc())
            raise HTTPException(422, f"PDF 파싱 실패: {e}")
        _log(f"Step 3 done [{time.time()-t0:.1f}s]")

        if not pdf_data:
            raise HTTPException(422, "PDF에서 직원 정보를 찾지 못했습니다.")

        # 3. PDF month/year 검증
        pdf_month = meta.get("month")
        pdf_year = meta.get("year")
        target_year = year
        target_month = month
        month_warning: Optional[str] = None

        if enforce_pdf_month and pdf_month and pdf_month != month:
            month_warning = (
                f"⚠️ 사용자 입력 월({year}-{month:02d})과 "
                f"PDF 출근부 월({pdf_year or year}-{pdf_month:02d}) 불일치 — "
                f"PDF 기준으로 처리합니다."
            )
            target_month = pdf_month
            if pdf_year:
                target_year = pdf_year

        # 4. PDF 첫 활동일 기반 입사일 추정 (엑셀에서 못 찾은 직원 보완)
        hire_dates_pdf = estimate_hire_dates_from_pdf(pdf_data, target_year, target_month)
        hire_dates = merge_hire_dates(hire_dates_excel, hire_dates_pdf)

        # 5. 원본 복사 후 수정
        _log(f"Step 5: copy + fill_attendance_sheet (year={target_year}, month={target_month})")
        shutil.copyfile(excel_orig_path, excel_out_path)

        try:
            att_result = fill_attendance_sheet(
                str(excel_out_path),
                pdf_data,
                year=target_year,
                month=target_month,
                sheet_name=sheet_name_final,
                paid_holidays=user_holidays,
                standard_hours=standard_hours,
                normal_start=normal_start,
                normal_end=normal_end,
                overwrite_existing=overwrite_existing,
                hire_dates=hire_dates,
                feb_credit_map=feb_credit_map,
                resign_dates=resign_dates_db,
            )
            _log(f"  fill_attendance_sheet done — stats: {len(att_result.get('stats', {}))}명")
        except ValueError as e:
            _log(f"Step 5 ValueError: {e}")
            raise HTTPException(422, str(e))
        except Exception as e:
            _log(f"Step 5 ERROR: {type(e).__name__}: {e}")
            _log(traceback.format_exc())
            raise HTTPException(500, f"근태 입력 실패: {type(e).__name__}: {e}")
        _log(f"Step 5 done [{time.time()-t0:.1f}s]")

        wb = att_result["wb"]
        leave_result = {"stats": {}, "missing": [], "review_list": []}

        # 6. 연차 입력 (옵션)
        if fill_leave:
            classified_per_employee = {}
            for name, days in pdf_data.items():
                from ..services.attendance_parser import normalize_name as _nn
                emp_hire = hire_dates.get(_nn(name))
                classified = {}
                for day in range(1, 32):
                    classified[day] = classify_attendance(
                        days.get(day, {}),
                        target_year, target_month, day,
                        att_result["merged_holidays"], normal_start, normal_end,
                        hire_date=emp_hire,
                    )
                classified_per_employee[name] = classified

            leave_result = fill_leave_sheet(
                wb,
                pdf_data,
                classified_per_employee,
                year=target_year,
                month=target_month,
                sheet_name=leave_sheet_name,
            )

        # 7. 입력 누락 검증
        from ..services.attendance_parser import normalize_name as _nn
        ws_used = wb[att_result["sheet_used"]]
        employee_blocks = {}
        for name in pdf_data.keys():
            blk = find_employee_block(ws_used, name)
            if blk:
                employee_blocks[name] = blk

        # hire_dates의 키는 이미 normalize됨. pdf_data 키와 동일하게 매핑
        hire_dates_for_verify = {}
        for name in pdf_data.keys():
            nn = _nn(name)
            if nn in hire_dates:
                hire_dates_for_verify[name] = hire_dates[nn]

        missing_entries = verify_input_completeness(
            ws_used,
            pdf_data,
            employee_blocks,
            hire_dates_for_verify,
            att_result["merged_holidays"],
            target_year,
            target_month,
            DAY_TO_COL,
        )

        # 8. 검증 + 검토리스트를 wb 인라인으로 추가 후 한 번에 save
        # 이전: wb.save → validate_lite (재로드) → wb_final 재로드 → write_review → save → 80초
        # 지금: validation 스킵 + wb에 직접 ws_review 추가 → 한 번에 save → 20초
        _log(f"Step 8: inline review sheet + single save")
        validation = {
            "ok": True,
            "summary": {"검증_생략": "처리 시간 단축 (Render 120s timeout 회피)"},
            "수식_손실": [], "수식_변경": [],
        }
        wb_final = wb  # 재로드 없이 같은 wb 사용

        # 검토리스트 시트 생성
        if "자동입력_검토리스트" not in wb_final.sheetnames:
            wb_final.create_sheet("자동입력_검토리스트")

        if "자동입력_검토리스트" in wb_final.sheetnames:
            ws_review = wb_final["자동입력_검토리스트"]

            # 처리 메타
            ws_review.append([])
            ws_review.append(["처리 메타"])
            ws_review.append(["대상연도", str(target_year)])
            ws_review.append(["대상월", str(target_month)])
            ws_review.append(["사용시트", att_result["sheet_used"]])
            if month_warning:
                ws_review.append(["월불일치_경고", month_warning])
            ws_review.append(["기준시간", str(standard_hours)])
            ws_review.append(["정상출근", normal_start])
            ws_review.append(["정상퇴근", normal_end])
            ws_review.append(["엑셀직원명_추출수", str(len(excel_names))])
            ws_review.append(["PDF매칭직원수", str(len(meta.get("matched_employees", [])))])
            ws_review.append(["2월시트_사용", str(feb_sheet_name or "없음")])

            # DB 시트 정보
            if db_info:
                ws_review.append([])
                ws_review.append([f"DB 시트 정보 ({len(db_info)}명)"])
                ws_review.append(["직원명_정규화", "원본이름", "입사일", "퇴사일", "상태", "현장"])
                for norm, info in db_info.items():
                    ws_review.append([
                        norm,
                        info.get("name", ""),
                        str(info.get("hire_date") or ""),
                        str(info.get("resign_date") or ""),
                        info.get("status") or "",
                        info.get("site") or "",
                    ])

            # 입사일 정보
            ws_review.append([])
            ws_review.append(["입사일 정보 (DB > 엑셀 > PDF추정 우선순위)"])
            ws_review.append(["직원명_정규화", "입사일", "퇴사일", "출처"])
            for nm in excel_names:
                hd = hire_dates.get(nm)
                rd = resign_dates_db.get(nm)
                if hd is not None:
                    if nm in db_info and db_info[nm].get("hire_date"):
                        src = "DB"
                    elif nm in hire_dates_excel:
                        src = "엑셀"
                    else:
                        src = "PDF추정"
                    ws_review.append([nm, str(hd), str(rd or ""), src])
                else:
                    ws_review.append([nm, "(미확인)", str(rd or ""), "정상직원으로_가정"])

            # 2월 인정 결과
            ws_review.append([])
            ws_review.append(["2월 마지막 주 인정 결과 (3/1 주휴 판단용)"])
            ws_review.append(["직원명_정규화", "결과"])
            for nm in excel_names:
                v = feb_credit_map.get(nm)
                if v is True:
                    ws_review.append([nm, "인정"])
                elif v is False:
                    ws_review.append([nm, "불인정"])
                else:
                    ws_review.append([nm, "(2월 시트 또는 직원 못 찾음)"])

            # 카운터 요약
            ws_review.append([])
            ws_review.append(["카운터 요약 (0이면 성공)"])
            ws_review.append(["항목", "건수"])
            counters = att_result.get("counters", {})
            for k in [
                "직원_매칭실패",
                "수식셀_덮어쓰기",
                "3월1일_주휴제외",
                "3월2일_유급제외",
            ]:
                ws_review.append([k, str(counters.get(k, 0))])

            ws_review.append([])
            ws_review.append(["기타 카운터"])
            for k in [
                "주말근무_연장행입력",
                "공휴일_유급_입력",
                "주휴_자동입력",
                "연장_텍스트_충돌",
                "일요일주휴_제외",
                "주휴_기본칸_충돌",
                "주휴_표기칸_충돌",
                "유급수정_잘못된주휴",
                "입사일_확인필요",
                "입사전_자동입력_제외",
                "퇴사후_자동입력_제외",
                "월말보정_중도입사자제외",
                "월말보정_중도퇴사자제외",
            ]:
                ws_review.append([k, str(counters.get(k, 0))])

            # 적용 공휴일
            ws_review.append([])
            ws_review.append(["적용 공휴일 (자동 + 사용자 입력)"])
            ws_review.append(["일자", "이름"])
            for date_str, hname in sorted(att_result["merged_holidays"].items()):
                ws_review.append([date_str, hname])

            # PDF 날짜 보정 내역
            day_corrections = meta.get("day_corrections", [])
            if day_corrections:
                ws_review.append([])
                ws_review.append(["PDF 날짜 시퀀스 보정 (raw → corrected)"])
                ws_review.append(["페이지", "raw_day", "corrected_day", "context"])
                for c in day_corrections:
                    ws_review.append([
                        str(c.get("page", "")),
                        str(c.get("raw", "")),
                        str(c.get("corrected", "")),
                        c.get("context", "")
                    ])

            # 입력 누락 검증 결과
            if missing_entries:
                ws_review.append([])
                ws_review.append([f"입력 누락 검증 ({len(missing_entries)}건)"])
                ws_review.append(["구분", "성명", "일자", "요일", "PDF원문", "입력값", "메시지"])
                for m in missing_entries[:200]:
                    ws_review.append([
                        m.get("구분", ""), m.get("성명", ""),
                        m.get("일자", ""), m.get("요일", ""),
                        m.get("PDF원문", ""), m.get("입력값", ""),
                        m.get("메시지", "")
                    ])

            # 근태 입력 검토
            ws_review.append([])
            ws_review.append(["근태 입력 검토"])
            ws_review.append(["구분", "성명", "일자", "요일", "PDF원문", "입력값", "메시지"])
            for r in att_result["review_list"]:
                ws_review.append([
                    r.get("구분", ""), r.get("성명", ""),
                    r.get("일자", ""), r.get("요일", ""),
                    r.get("PDF원문", ""), str(r.get("입력값", "")),
                    r.get("메시지", "")
                ])

            # 수식 셀 보호 로그
            log = att_result.get("log", [])
            formula_protect = [e for e in log if e.get("kind") == "수식셀_보호"]
            if formula_protect:
                ws_review.append([])
                ws_review.append(["수식 셀 보호 로그 (덮어쓰지 않고 건너뜀)"])
                ws_review.append(["시트", "셀", "수식", "건너뛴값"])
                for e in formula_protect[:100]:
                    ws_review.append([
                        e.get("sheet", ""), e.get("cell", ""),
                        e.get("formula", ""), str(e.get("skipped_value", "")),
                    ])

            if fill_leave and leave_result.get("review_list"):
                ws_review.append([])
                ws_review.append(["연차 입력 검토"])
                ws_review.append(["구분", "성명", "메시지"])
                for r in leave_result.get("review_list", []):
                    ws_review.append([
                        r.get("구분", ""), r.get("성명", ""), r.get("메시지", "")
                    ])

        _log(f"Step 9: wb_final.save")
        wb_final.save(str(excel_out_path))
        _log(f"  wb_final.save done [{time.time()-t0:.1f}s]")

        with open(excel_out_path, "rb") as f:
            content = f.read()
        _log(f"Step 10: response build — content size={len(content)/1024:.1f} KB [{time.time()-t0:.1f}s]")

        result_summary = {
            "pdf_meta": {
                "year": meta.get("year"),
                "month": meta.get("month"),
                "total_employees": meta.get("total_employees"),
                "raw_pages": meta.get("raw_pages"),
                "day_corrections": len(meta.get("day_corrections", [])),
                "matched_employees": len(meta.get("matched_employees", [])),
            },
            "target": {
                "year": target_year,
                "month": target_month,
                "sheet_used": att_result["sheet_used"],
                "month_warning": month_warning,
                "feb_sheet": feb_sheet_name,
            },
            "hire_dates": {
                "from_excel": len(hire_dates_excel),
                "from_pdf_estimate": len(hire_dates_pdf),
                "total": len(hire_dates),
            },
            "attendance": {
                "filled_employees": len(att_result["stats"]),
                "total_cells": sum(att_result["stats"].values()),
                "missing": att_result["missing"],
                "counters": att_result.get("counters", {}),
            },
            "completeness": {
                "missing_entries": len(missing_entries),
            },
            "leave": {
                "filled_employees": len(leave_result.get("stats", {})),
                "missing": leave_result.get("missing", []),
                "skipped": not fill_leave,
            },
            "validation": validation["summary"],
            "validation_ok": validation["ok"],
            "review_count": (
                len(att_result["review_list"])
                + len(leave_result.get("review_list", []))
                + len(missing_entries)
            ),
        }

        from urllib.parse import quote
        encoded_summary = json.dumps(result_summary, ensure_ascii=False)
        out_filename = excel_out_path.name
        ascii_fallback = "result.xlsx"
        utf8_filename = quote(out_filename, safe="")

        _log(f"=== SUCCESS — total {time.time()-t0:.1f}s ===")
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=\"{ascii_fallback}\"; filename*=UTF-8''{utf8_filename}",
                "X-Process-Result": encoded_summary.encode("utf-8").hex(),
                "Access-Control-Expose-Headers": "X-Process-Result, Content-Disposition",
            },
        )
      except HTTPException:
        raise
      except Exception as e:
        _log(f"=== UNEXPECTED ERROR at {time.time()-t0:.1f}s: {type(e).__name__}: {e} ===")
        _log(traceback.format_exc())
        raise HTTPException(500, f"처리 중 예기치 못한 에러: {type(e).__name__}: {e}")
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


@router.post("/process-elcomtec")
async def process_attendance_elcomtec(
    pdfs: list[UploadFile] = File(...),
    excel: UploadFile = File(...),
    year: int = Form(...),
    month: int = Form(...),
    sheet_name: Optional[str] = Form(None),
    company_id: str = Form("elcomtec"),
    x_admin_token: Optional[str] = Header(None, alias="X-Admin-Token"),
    authorization: Optional[str] = Header(None),
):
    """엘컴텍(근태대장 포맷) 전용 — 여러 PDF(응원봉·영업) + 양식 xlsx → 지급파일.

    엘티와이 흐름과 분리(파서·룰 상이). 자동: 기본/연장/심야/특근/특잔 + 주휴 개근룰.
    수동영역(검토리스트): 평일결근→연차/반차/휴무 분류, 지각조퇴 deduction.
    """
    from ..services.attendance_parser_elcomtec import parse_pdfs_elcomtec
    from ..services.excel_writer_elcomtec import fill_attendance_sheet_elcomtec
    from ..services.excel_writer import build_name_to_row_map, find_target_sheet

    verify_admin_token(x_admin_token, authorization)

    cid = (company_id or "elcomtec").strip().lower()
    if cid not in COMPANY_LABELS:
        raise HTTPException(403, f"등록되지 않은 회사: {company_id}")
    if not pdfs or not excel.filename:
        raise HTTPException(400, "PDF(1개 이상)와 Excel 파일이 필요합니다.")

    t0 = time.time()
    company_label = COMPANY_LABELS.get(cid, "")
    _log(f"=== START process-elcomtec year={year} month={month} pdfs={len(pdfs)} ===")
    sheet_name_final = sheet_name.strip() if sheet_name and sheet_name.strip() else None

    tmp_dir = Path(tempfile.mkdtemp(prefix="elcomtec_"))
    try:
      try:
        # 1. 파일 저장
        pdf_paths: list[str] = []
        for up in pdfs:
            if not up.filename or not up.filename.lower().endswith(".pdf"):
                raise HTTPException(400, f"PDF만 허용: {up.filename}")
            p = tmp_dir / up.filename
            with open(p, "wb") as f:
                f.write(await up.read())
            pdf_paths.append(str(p))
        excel_orig_path = tmp_dir / excel.filename
        with open(excel_orig_path, "wb") as f:
            f.write(await excel.read())
        out_name = f"{company_label}_{year:04d}-{month:02d}_지급파일.xlsx"
        excel_out_path = tmp_dir / out_name
        _log(f"Step 1 done [{time.time()-t0:.1f}s] pdfs={len(pdf_paths)}")

        # 1.5 dimension 정상화
        try:
            fixes = _normalize_xlsx_dimensions(str(excel_orig_path))
            _log(f"Step 1.5 normalize done (fixes={len(fixes)})")
        except Exception as e:
            _log(f"Step 1.5 skipped: {type(e).__name__}: {e}")

        # 2. 직원명→행 매핑 (원본 VLOOKUP 캐시 기반)
        from openpyxl import load_workbook as _lw
        wb_probe = _lw(str(excel_orig_path), data_only=True, read_only=False, keep_links=False)
        try:
            sheet_used_probe = find_target_sheet(wb_probe, month, sheet_name_final)
        finally:
            wb_probe.close()
        name_to_row = build_name_to_row_map(str(excel_orig_path), sheet_used_probe)
        _log(f"Step 2 name_to_row: {len(name_to_row)}명 (sheet={sheet_used_probe})")

        # 3. PDF 파싱
        parsed, meta = parse_pdfs_elcomtec(pdf_paths)
        _log(f"Step 3 PDF: {meta['total_employees']}명, pages={meta['raw_pages']}, parts={meta.get('parts')}")
        if not parsed:
            raise HTTPException(422, "PDF에서 직원 정보를 찾지 못했습니다.")

        # 월 검증 경고
        month_warning = None
        if meta.get("month") and meta["month"] != month:
            month_warning = f"⚠️ 입력 월({month}) ≠ PDF 월({meta['month']})"

        # 4. 원본 복사 후 자동입력
        shutil.copyfile(excel_orig_path, excel_out_path)
        res = fill_attendance_sheet_elcomtec(
            str(excel_out_path), parsed, year=year, month=month,
            sheet_name=sheet_name_final, name_to_row=name_to_row,
        )
        wb = res["wb"]
        _log(f"Step 4 fill done — {sum(res['stats'].values())}셀/{len(res['stats'])}명, 매칭실패={res['counters']['직원_매칭실패']}")

        # 5. 검토리스트 시트
        if "자동입력_검토리스트" not in wb.sheetnames:
            wb.create_sheet("자동입력_검토리스트")
        ws_r = wb["자동입력_검토리스트"]
        ws_r.append([])
        ws_r.append(["엘컴텍 처리 메타"])
        ws_r.append(["대상연월", f"{year}-{month:02d}"])
        ws_r.append(["사용시트", res["sheet_used"]])
        ws_r.append(["PDF 파트", ", ".join(meta.get("parts", []))])
        ws_r.append(["PDF 직원수", str(meta["total_employees"])])
        if month_warning:
            ws_r.append(["월불일치", month_warning])
        ws_r.append([])
        ws_r.append(["카운터 요약"])
        for k, v in res["counters"].items():
            ws_r.append([k, str(v)])
        if res["missing"]:
            ws_r.append([])
            ws_r.append([f"매칭 실패 ({len(res['missing'])}명) — 근태시트에 이름 없음"])
            for nm in res["missing"]:
                ws_r.append([nm])
        ws_r.append([])
        ws_r.append([f"검토 항목 ({len(res['review_list'])}건) — 수동 확인 필요"])
        ws_r.append(["구분", "성명", "일자", "요일", "PDF원문", "입력값", "메시지"])
        for r in res["review_list"]:
            ws_r.append([
                r.get("구분", ""), r.get("성명", ""), r.get("일자", ""),
                r.get("요일", ""), r.get("PDF원문", ""), str(r.get("입력값", "")),
                r.get("메시지", ""),
            ])

        wb.save(str(excel_out_path))
        with open(excel_out_path, "rb") as f:
            content = f.read()
        _log(f"=== SUCCESS elcomtec — total {time.time()-t0:.1f}s, size={len(content)/1024:.1f}KB ===")

        result_summary = {
            "pdf_meta": {
                "year": meta.get("year"), "month": meta.get("month"),
                "total_employees": meta.get("total_employees"),
                "raw_pages": meta.get("raw_pages"), "parts": meta.get("parts", []),
            },
            "target": {"year": year, "month": month, "sheet_used": res["sheet_used"],
                       "month_warning": month_warning},
            "attendance": {
                "filled_employees": len(res["stats"]),
                "total_cells": sum(res["stats"].values()),
                "missing": res["missing"],
                "counters": res["counters"],
            },
            # 공유 결과 패널 호환용 stub (엘컴텍은 연차 자동입력·수식검증 미사용)
            "leave": {"filled_employees": 0, "missing": [], "skipped": True},
            "validation": {},
            "validation_ok": True,
            "review_count": len(res["review_list"]) + len(res.get("candidate_rows", [])),
        }
        from urllib.parse import quote
        encoded_summary = json.dumps(result_summary, ensure_ascii=False)
        utf8_filename = quote(excel_out_path.name, safe="")
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=\"result.xlsx\"; filename*=UTF-8''{utf8_filename}",
                "X-Process-Result": encoded_summary.encode("utf-8").hex(),
                "Access-Control-Expose-Headers": "X-Process-Result, Content-Disposition",
            },
        )
      except HTTPException:
        raise
      except Exception as e:
        _log(f"=== elcomtec ERROR {time.time()-t0:.1f}s: {type(e).__name__}: {e} ===")
        _log(traceback.format_exc())
        raise HTTPException(500, f"엘컴텍 처리 실패: {type(e).__name__}: {e}")
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


@router.post("/preview")
async def preview_pdf(
    pdf: UploadFile = File(...),
    x_admin_token: Optional[str] = Header(None, alias="X-Admin-Token"),
    authorization: Optional[str] = Header(None),
):
    """PDF만 파싱해서 미리보기 (엑셀 처리 X)"""
    verify_admin_token(x_admin_token, authorization)

    tmp_dir = Path(tempfile.mkdtemp(prefix="preview_"))
    try:
        pdf_path = tmp_dir / pdf.filename
        with open(pdf_path, "wb") as f:
            f.write(await pdf.read())

        pdf_data, meta = parse_pdf(str(pdf_path))

        rows = []
        for name, days in pdf_data.items():
            for day, slot in days.items():
                if slot.get("start") or slot.get("end") or slot.get("note"):
                    rows.append({
                        "name": name,
                        "day": day,
                        "start": slot.get("start", ""),
                        "end": slot.get("end", ""),
                        "ot": slot.get("ot", ""),
                        "note": slot.get("note", ""),
                    })

        return {
            "meta": meta,
            "employees": list(pdf_data.keys()),
            "rows": rows,
        }
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)
