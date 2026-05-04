"""
저장 후 검증 — 원본과 결과 파일의 수식/서식 보존 여부 비교
"""
from openpyxl import load_workbook


def collect_formulas(wb) -> dict:
    """워크북의 모든 수식 셀 수집"""
    formulas = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    formulas[f"{ws.title}!{cell.coordinate}"] = v
    return formulas


def collect_merges(wb) -> dict:
    """병합셀 범위 수집"""
    merges = {}
    for ws in wb.worksheets:
        merges[ws.title] = sorted(str(m) for m in ws.merged_cells.ranges)
    return merges


def collect_dimensions(wb) -> dict:
    """행 높이 / 열 너비 수집"""
    dims = {}
    for ws in wb.worksheets:
        rows = {r: dim.height for r, dim in ws.row_dimensions.items() if dim.height}
        cols = {c: dim.width for c, dim in ws.column_dimensions.items() if dim.width}
        dims[ws.title] = {"rows": rows, "cols": cols}
    return dims


def validate(original_path: str, output_path: str) -> dict:
    """
    원본 vs 결과 비교.
    Returns: { ok, formula_diff, merge_diff, dim_diff, summary }
    """
    wb_orig = load_workbook(original_path, data_only=False)
    wb_new = load_workbook(output_path, data_only=False)

    # 시트명 비교
    orig_sheets = set(wb_orig.sheetnames)
    new_sheets = set(wb_new.sheetnames)
    sheet_added = new_sheets - orig_sheets
    sheet_removed = orig_sheets - new_sheets

    # 수식 비교
    f_orig = collect_formulas(wb_orig)
    f_new = collect_formulas(wb_new)

    formula_lost = {k: v for k, v in f_orig.items() if k not in f_new}
    formula_changed = {
        k: {"orig": f_orig[k], "new": f_new[k]}
        for k in f_orig if k in f_new and f_orig[k] != f_new[k]
    }
    formula_added = {k: v for k, v in f_new.items() if k not in f_orig}

    # 병합셀 비교
    m_orig = collect_merges(wb_orig)
    m_new = collect_merges(wb_new)
    merge_diff = {}
    for sheet in orig_sheets:
        if sheet in new_sheets:
            o = set(m_orig.get(sheet, []))
            n = set(m_new.get(sheet, []))
            if o != n:
                merge_diff[sheet] = {
                    "lost": sorted(o - n),
                    "added": sorted(n - o),
                }

    # 행/열 크기 비교 (간단)
    d_orig = collect_dimensions(wb_orig)
    d_new = collect_dimensions(wb_new)
    dim_diff = {}
    for sheet in orig_sheets:
        if sheet in new_sheets:
            ro = d_orig.get(sheet, {}).get("rows", {})
            rn = d_new.get(sheet, {}).get("rows", {})
            co = d_orig.get(sheet, {}).get("cols", {})
            cn = d_new.get(sheet, {}).get("cols", {})
            row_changed = sum(1 for k in ro if rn.get(k) != ro[k])
            col_changed = sum(1 for k in co if cn.get(k) != co[k])
            if row_changed or col_changed:
                dim_diff[sheet] = {
                    "row_changed": row_changed,
                    "col_changed": col_changed,
                }

    summary = {
        "원본_수식_개수": len(f_orig),
        "결과_수식_개수": len(f_new),
        "수식_손실": len(formula_lost),
        "수식_변경": len(formula_changed),
        "병합셀_변경된_시트": len(merge_diff),
        "치수_변경된_시트": len(dim_diff),
        "추가된_시트": list(sheet_added),
        "삭제된_시트": list(sheet_removed),
    }

    ok = (
        len(formula_lost) == 0 and
        len(formula_changed) == 0 and
        len(sheet_removed) == 0 and
        len(merge_diff) == 0
    )

    return {
        "ok": ok,
        "summary": summary,
        "formula_lost": formula_lost,
        "formula_changed": formula_changed,
        "formula_added": formula_added,
        "merge_diff": merge_diff,
        "dim_diff": dim_diff,
    }


def write_validation_to_review(wb, validation: dict):
    """검증 결과를 워크북의 '자동입력_검토리스트' 시트에 추가"""
    sheet_name = "자동입력_검토리스트"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    ws.append(["검증 결과"])
    ws.append([f"전체 OK: {validation['ok']}"])
    ws.append([])
    ws.append(["항목", "값"])
    for k, v in validation["summary"].items():
        ws.append([k, str(v)])

    if validation["formula_lost"]:
        ws.append([])
        ws.append(["손실된 수식", ""])
        ws.append(["위치", "수식"])
        for loc, formula in list(validation["formula_lost"].items())[:50]:
            ws.append([loc, formula])

    if validation["formula_changed"]:
        ws.append([])
        ws.append(["변경된 수식", ""])
        ws.append(["위치", "원본 → 결과"])
        for loc, diff in list(validation["formula_changed"].items())[:50]:
            ws.append([loc, f"{diff['orig']} → {diff['new']}"])
