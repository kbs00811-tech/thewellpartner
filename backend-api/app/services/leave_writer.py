"""
연차내역 시트 자동 입력 (수식/서식 보존)

엘티와이 양식:
- 직원당 묶음 단위 (발생/사용/잔여/입사/퇴사/계 6행)
- C/D열: 직원명
- E열: 발생/사용/잔여
- 4월(7번 컬럼) 사용 칸에 0.5/1/2.5 등 입력
"""
from __future__ import annotations
from typing import Dict, List
import datetime
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell

from .attendance_parser import normalize_name
from .excel_writer import safe_set_value, is_formula_cell


def find_leave_employee_block(ws, employee_name: str):
    """
    연차내역 시트에서 직원의 시작 행 + 발생/사용/잔여 행 위치 찾기.
    Returns: { "start_row", "발생_row", "사용_row", "잔여_row" } or None
    """
    name_norm = normalize_name(employee_name)
    for r in range(1, ws.max_row + 1):
        c_val = ws.cell(row=r, column=3).value or ws.cell(row=r, column=4).value
        e_val = ws.cell(row=r, column=5).value
        if c_val and normalize_name(str(c_val)) == name_norm and \
           isinstance(e_val, str) and e_val.strip() == "발생":
            return {
                "start_row": r,
                "발생_row": r,
                "사용_row": r + 1,
                "잔여_row": r + 2,
            }
    return None


def get_target_month_col(ws, target_month: int) -> int | None:
    """
    연차내역 시트 헤더에서 대상 월의 컬럼 찾기.
    헤더 행에 "4월", "5월"... 형태로 표시됨.
    """
    target_label = f"{target_month}월"
    # 처음 8행 검색
    for r in range(1, 8):
        for c in range(5, 35):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip() == target_label:
                return c
    return None


def collect_leave_usage(employee_data: Dict[int, dict],
                        year: int, month: int) -> tuple[float, str]:
    """
    PDF 데이터에서 연차/반차/반반차 사용량 + 메모 생성.
    Returns: (total_count, memo_string)
    예: (2.5, "3/3(화) 연차, 3/16(월) 반차, 3/26(목) 반반차")
    """
    items: List[tuple[int, str, str]] = []  # (day, dow, type)
    total = 0.0

    for day in range(1, 32):
        slot = employee_data.get(day, {})
        note = slot.get("note", "")
        if note in ("연차",):
            count = 1.0
        elif note in ("반차",):
            count = 0.5
        elif note in ("반반차", "반반"):
            count = 0.25
        else:
            continue

        try:
            d = datetime.date(year, month, day)
            dow = ["월", "화", "수", "목", "금", "토", "일"][d.weekday()]
        except ValueError:
            continue

        items.append((day, dow, note if note != "반반" else "반반차"))
        total += count

    memo = ", ".join(f"{month}/{d}({w}) {t}" for d, w, t in items)
    return total, memo


def is_full_attendance(employee_classified: Dict[int, dict],
                       year: int, month: int) -> bool:
    """
    만근 판단 — 해당 월 평일 + 토요일 근무대상일에 대해
    출근/연차/반차/반반차/유급 등 인정 항목 있으면 만근.
    결근(평일에 빈 슬롯)이 1건이라도 있으면 만근 아님.
    """
    try:
        from calendar import monthrange
        last_day = monthrange(year, month)[1]
    except Exception:
        last_day = 31

    for day in range(1, last_day + 1):
        try:
            d = datetime.date(year, month, day)
        except ValueError:
            continue
        if d.weekday() == 6:  # 일요일은 만근 판단에서 제외
            continue

        c = employee_classified.get(day, {})
        # 인정 조건
        recognized = (
            c.get("기본", 0) >= 4 or
            c.get("특근", 0) > 0 or
            c.get("비고") in ("연차", "반차", "반반차", "유급")
        )
        if not recognized:
            return False
    return True


def fill_leave_sheet(
    wb: Workbook,
    pdf_data: Dict[str, Dict[int, dict]],
    classified_per_employee: Dict[str, Dict[int, dict]],
    year: int,
    month: int,
    sheet_name: str = "연차내역",
) -> dict:
    """
    연차내역 시트에 발생/사용/메모 자동 입력.
    원본 wb를 그대로 받아 in-place 수정.
    """
    if sheet_name not in wb.sheetnames:
        return {"error": f"연차내역 시트 없음. 사용 가능: {wb.sheetnames}"}

    ws = wb[sheet_name]
    log = []
    review = []
    stats = {}
    missing = []

    # 대상 월 컬럼 찾기
    target_col = get_target_month_col(ws, month)
    if target_col is None:
        review.append({
            "구분": "월컬럼_미발견",
            "메시지": f"{month}월 컬럼을 헤더에서 찾지 못함. 헤더 확인 필요",
        })
        # 폴백: 4월 = 7번 컬럼 (엘티와이 양식 기본)
        target_col = 6 + month  # G(7)부터 1월? 양식별 다름
        # 안전하게: 메모는 G열, 사용은 다른 곳일 수 있음
        # 검토리스트에 남기고 진행

    # 메모 컬럼 — 대개 마지막 컬럼 또는 별도 비고 칸
    # 우선 target_col 옆 또는 별도 지정 가능. 일단 사용 컬럼만 입력
    memo_col = None  # 양식에서 별도 메모 컬럼 찾기
    for r in range(1, 8):
        for c in range(target_col + 1, target_col + 5):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and ("메모" in v or "비고" in v):
                memo_col = c
                break
        if memo_col:
            break

    for name, days in pdf_data.items():
        block = find_leave_employee_block(ws, name)
        if not block:
            missing.append(name)
            review.append({
                "구분": "연차직원_매칭실패",
                "성명": name,
                "메시지": "연차내역 시트에서 행 못 찾음",
            })
            continue

        # 사용량 + 메모
        usage, memo = collect_leave_usage(days, year, month)

        # 발생 (만근 시 1)
        is_full = is_full_attendance(classified_per_employee.get(name, {}), year, month)
        발생값 = 1 if is_full else 0

        cell_count = 0

        # 발생 입력
        if safe_set_value(ws, block["발생_row"], target_col, 발생값, log):
            cell_count += 1

        # 사용 입력
        if usage > 0:
            if safe_set_value(ws, block["사용_row"], target_col, usage, log):
                cell_count += 1

        # 메모
        if memo and memo_col:
            if safe_set_value(ws, block["사용_row"], memo_col, memo, log):
                cell_count += 1
        elif memo:
            # 메모 컬럼 못 찾았을 때 검토리스트에만
            review.append({
                "구분": "연차메모",
                "성명": name,
                "메시지": memo,
            })

        stats[name] = cell_count
        review.append({
            "구분": "연차_입력",
            "성명": name,
            "메시지": f"발생 {발생값}, 사용 {usage}, 메모: {memo or '없음'}",
        })

    return {
        "stats": stats,
        "missing": missing,
        "log": log,
        "review_list": review,
    }
