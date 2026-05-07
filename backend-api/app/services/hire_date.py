"""
직원별 입사일 추출 + 2월 근태 시트 인정 여부 검사

입사일 우선순위:
  1. 엑셀 시트의 직원 블록 근처(B/C/E/G열, 6행 안)에 있는 날짜 셀
  2. PDF 데이터에서 첫 근태 활동일 (첫째 주 안이 아니면 중도입사 추정)
  3. 못 찾으면 None — 호출자가 검토리스트 처리
"""
from __future__ import annotations
import re
import datetime
from calendar import monthrange
from typing import Dict, List, Optional, Tuple
from openpyxl.workbook.workbook import Workbook


# 날짜 파싱 정규식 (4자리 연도 우선)
_DATE_PATTERNS_4Y = re.compile(r"(\d{4})[-./\s년](\d{1,2})[-./\s월](\d{1,2})")
_DATE_PATTERNS_2Y = re.compile(r"(\d{2})[-./](\d{1,2})[-./](\d{1,2})")


def try_parse_date(text: str) -> Optional[datetime.date]:
    """문자열에서 날짜 파싱. YYYY-MM-DD / YY.MM.DD / YYYY.M.D 등 다양한 포맷."""
    if not text:
        return None
    s = str(text).strip()
    if not s:
        return None

    m = _DATE_PATTERNS_4Y.search(s)
    if m:
        try:
            return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except (ValueError, IndexError):
            pass

    m = _DATE_PATTERNS_2Y.search(s)
    if m:
        try:
            yr = int(m.group(1))
            year = 2000 + yr if yr < 80 else 1900 + yr
            return datetime.date(year, int(m.group(2)), int(m.group(3)))
        except (ValueError, IndexError):
            pass

    return None


def _looks_like_hire_date(d: datetime.date) -> bool:
    """입사일로 합리적인 날짜인지 (1990~2030)"""
    return 1990 <= d.year <= 2030


def _normalize_name(name: str) -> str:
    return str(name or "").replace(" ", "").strip()


def extract_hire_dates_from_excel(
    wb: Workbook,
    sheet_names: List[str],
    expected_names: List[str],
) -> Dict[str, datetime.date]:
    """
    엑셀 시트들에서 직원별 입사일 추출 시도.

    검색 범위:
      - 직원 블록 시작 행 ± 0~5행
      - 1~7번 열 (A~G) 안의 날짜 셀

    Args:
      wb: 워크북
      sheet_names: 검색할 시트 이름 목록 (근태 (3월), 직원정보 등)
      expected_names: 매칭 대상 직원명

    Returns:
      {정규화_이름: 입사일}
    """
    BASIC_LABELS = {"기본", "기본근무", "정상"}
    hire_dates: Dict[str, datetime.date] = {}

    expected_set = {_normalize_name(n) for n in expected_names if n}

    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        for r in range(1, ws.max_row + 1):
            d_val = ws.cell(row=r, column=4).value
            f_val = ws.cell(row=r, column=6).value
            if not (d_val and isinstance(d_val, str)):
                continue
            label = str(f_val).strip() if f_val else ""
            if label not in BASIC_LABELS:
                continue

            name = _normalize_name(d_val)
            if not name or name in hire_dates:
                continue
            if expected_set and name not in expected_set:
                continue

            # 직원 블록 시작 행 + 0~5행 / 1~7번 열 검사
            for row_offset in range(0, 6):
                for col in range(1, 8):
                    if col == 6:  # F열은 라벨이라 스킵
                        continue
                    cell = ws.cell(row=r + row_offset, column=col)
                    v = cell.value
                    parsed: Optional[datetime.date] = None
                    if isinstance(v, datetime.datetime):
                        parsed = v.date()
                    elif isinstance(v, datetime.date):
                        parsed = v
                    elif isinstance(v, str):
                        parsed = try_parse_date(v)
                    if parsed and _looks_like_hire_date(parsed):
                        hire_dates[name] = parsed
                        break
                if name in hire_dates:
                    break

    return hire_dates


def estimate_hire_dates_from_pdf(
    pdf_data: Dict[str, Dict[int, dict]],
    year: int,
    month: int,
) -> Dict[str, datetime.date]:
    """
    PDF 데이터에서 직원별 첫 활동일을 입사일로 추정.

    규칙:
      - 첫 활동일이 첫째 주(1~7일) 안이면 정상 직원 → 추정 안 함
      - 첫째 주 밖이면 그 날을 입사일로 추정 (중도입사자)

    Returns:
      {정규화_이름: 추정_입사일}  (정상 직원은 미포함)
    """
    estimated: Dict[str, datetime.date] = {}

    for name, days in pdf_data.items():
        norm = _normalize_name(name)
        if not norm:
            continue

        first_active_day: Optional[int] = None
        for day in sorted(days.keys()):
            slot = days[day]
            has_data = bool(
                slot.get("start") or slot.get("end") or slot.get("ot")
                or slot.get("note") in ("연차", "반차", "반반차", "유급", "주휴")
            )
            if has_data:
                first_active_day = day
                break

        if first_active_day is None:
            # 활동 데이터 없음 → 추정 못 함
            continue

        if first_active_day <= 7:
            # 첫째 주 안 → 정상 직원
            continue

        # 첫째 주 밖 → 중도입사 추정
        try:
            estimated[norm] = datetime.date(year, month, first_active_day)
        except ValueError:
            continue

    return estimated


def merge_hire_dates(
    primary: Dict[str, datetime.date],
    secondary: Dict[str, datetime.date],
) -> Dict[str, datetime.date]:
    """
    primary 우선 (엑셀 추출), secondary 보완 (PDF 추정).
    """
    merged = dict(secondary)
    merged.update(primary)
    return merged


def extract_employee_info_from_db(
    wb: Workbook,
    db_sheet_candidates: Optional[List[str]] = None,
) -> Dict[str, Dict]:
    """
    DB 시트에서 직원별 정보 추출 (입사일/퇴사일/재직여부).

    헤더에서 컬럼명 자동 감지 (성명/이름, 입사일, 퇴사일, 재직여부 등).

    Args:
      wb: 워크북
      db_sheet_candidates: 검색할 시트 이름 목록 (기본: ["DB", "직원정보", "사원정보"])

    Returns:
      { 정규화_이름: {
            "name": str (원본),
            "hire_date": date 또는 None,
            "resign_date": date 또는 None,
            "status": str 또는 None,
            "site": str 또는 None,
        }
      }
    """
    if db_sheet_candidates is None:
        db_sheet_candidates = ["DB", "db", "직원정보", "사원정보", "직원DB", "근무자정보"]

    NAME_LABELS = {"성명", "이름", "근무자명", "직원명", "사원명", "name", "Name"}
    HIRE_LABELS = {"입사일", "입사일자", "입사", "입사년월일", "입사 일자", "hire", "hire_date"}
    RESIGN_LABELS = {"퇴사일", "퇴사일자", "퇴사", "퇴사년월일", "퇴사 일자", "resign", "resign_date"}
    STATUS_LABELS = {"재직", "재직여부", "상태", "근무상태", "근무 상태", "재직 여부", "status"}
    SITE_LABELS = {"업체", "현장", "현장명", "업체명", "근무지", "사이트"}

    info_map: Dict[str, Dict] = {}

    target_sheet: Optional[str] = None
    for cand in db_sheet_candidates:
        if cand in wb.sheetnames:
            target_sheet = cand
            break

    if target_sheet is None:
        return info_map

    ws = wb[target_sheet]

    # 헤더 행 자동 감지 (1~5행 검사)
    header_row: Optional[int] = None
    name_col = hire_col = resign_col = status_col = site_col = None

    for r in range(1, min(8, ws.max_row + 1)):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if not v or not isinstance(v, str):
                continue
            vs = v.strip()
            if vs in NAME_LABELS and name_col is None:
                name_col = c
                header_row = r
            elif vs in HIRE_LABELS and hire_col is None:
                hire_col = c
            elif vs in RESIGN_LABELS and resign_col is None:
                resign_col = c
            elif vs in STATUS_LABELS and status_col is None:
                status_col = c
            elif vs in SITE_LABELS and site_col is None:
                site_col = c
        if name_col is not None and (hire_col or resign_col):
            break

    if name_col is None or header_row is None:
        return info_map

    # 데이터 행 순회
    for r in range(header_row + 1, ws.max_row + 1):
        name_val = ws.cell(row=r, column=name_col).value
        if not name_val:
            continue
        name_str = str(name_val).strip()
        if not name_str:
            continue
        norm = _normalize_name(name_str)
        if not norm or norm in info_map:
            continue

        info: Dict = {
            "name": name_str,
            "hire_date": None,
            "resign_date": None,
            "status": None,
            "site": None,
        }

        if hire_col:
            hv = ws.cell(row=r, column=hire_col).value
            info["hire_date"] = _parse_date_value(hv)
        if resign_col:
            rv = ws.cell(row=r, column=resign_col).value
            info["resign_date"] = _parse_date_value(rv)
        if status_col:
            sv = ws.cell(row=r, column=status_col).value
            if sv:
                info["status"] = str(sv).strip()
        if site_col:
            stv = ws.cell(row=r, column=site_col).value
            if stv:
                info["site"] = str(stv).strip()

        info_map[norm] = info

    return info_map


def _parse_date_value(v) -> Optional[datetime.date]:
    """셀 값을 날짜로 파싱"""
    if v is None or v == "":
        return None
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    if isinstance(v, str):
        return try_parse_date(v)
    return None


def find_feb_attendance_sheet(wb: Workbook, year: int, march_month: int = 3) -> Optional[str]:
    """근태 (2월) 시트 자동 감지"""
    feb_month = march_month - 1
    if feb_month < 1:
        return None
    patterns = [
        rf"근태\s*\(\s*{feb_month}\s*월\s*\)",
        rf"근태\s*{feb_month}\s*월",
        rf"{feb_month}\s*월\s*근태",
    ]
    for sn in wb.sheetnames:
        for p in patterns:
            if re.search(p, sn):
                return sn
    return None


def check_feb_last_week_credit(
    wb: Workbook,
    feb_sheet_name: str,
    employee_name: str,
    year: int,
    feb_month: int = 2,
    day_to_col: Optional[Dict[int, int]] = None,
) -> Optional[bool]:
    """
    2월 마지막 주 평일(월~금) 근태 인정 여부 검사.

    Returns:
      True  — 마지막 주 5일 모두 인정
      False — 일부 또는 전부 불인정
      None  — 시트 없음 / 직원 못 찾음 (판단 불가)
    """
    BASIC_LABELS = {"기본", "기본근무", "정상"}
    if feb_sheet_name not in wb.sheetnames:
        return None
    ws = wb[feb_sheet_name]

    if day_to_col is None:
        day_to_col = {day: 7 + day for day in range(1, 32)}

    # 직원 블록 찾기
    name_norm = _normalize_name(employee_name)
    block_start = None
    for r in range(1, ws.max_row + 1):
        d_val = ws.cell(row=r, column=4).value
        f_val = ws.cell(row=r, column=6).value
        if not (d_val and _normalize_name(str(d_val)) == name_norm):
            continue
        if isinstance(f_val, str) and f_val.strip() in BASIC_LABELS:
            block_start = r
            break

    if block_start is None:
        return None

    basic_row = block_start
    overtime_row = block_start + 1

    # 2월 마지막 주 평일 (월~금)
    last_day_of_feb = monthrange(year, feb_month)[1]  # 28 or 29
    # 마지막 금요일 찾기
    last_friday = None
    for d in range(last_day_of_feb, 0, -1):
        try:
            if datetime.date(year, feb_month, d).weekday() == 4:
                last_friday = d
                break
        except ValueError:
            continue
    if last_friday is None:
        return None

    last_monday = last_friday - 4
    if last_monday < 1:
        return None

    credit_set = {"연차", "반차", "반반차", "유급"}

    for day in range(last_monday, last_friday + 1):
        col = day_to_col.get(day)
        if not col:
            return None
        basic_val = ws.cell(row=basic_row, column=col).value
        note_val = ws.cell(row=overtime_row, column=col).value

        is_credit = False
        if basic_val in (8, 8.0):
            is_credit = True
        elif isinstance(basic_val, (int, float)) and basic_val >= 8:
            is_credit = True
        if isinstance(note_val, str) and note_val.strip() in credit_set:
            is_credit = True

        if not is_credit:
            return False

    return True
