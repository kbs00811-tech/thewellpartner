"""
근태 시트 자동 입력 (수식/서식 보존)

엘티와이 양식 — 직원당 6행 구조:
  1. 기본
  2. 연장   ← 시간(숫자) + 텍스트(연차/반차/반반차/유급/주휴) 동시 담당
  3. 심야
  4. 특근
  5. 특잔
  6. 지각 조퇴

비고 행은 별도로 존재하지 않음.
유급/연차/반차/반반차/주휴 텍스트는 모두 연장 행(basic_row + 1)에 입력.

원칙:
  1. 원본 파일은 shutil.copyfile로 복사
  2. load_workbook(..., data_only=False) 사용
  3. 수식 셀 절대 덮어쓰지 않음
  4. 병합셀 좌상단만 입력
  5. 입력은 H~AL열 (1일~31일)만
  6. 토요일/일요일 근무는 연장 행에 숫자 입력 (특근 행 X)
  7. 평일 실제 연장 숫자가 있는 날에는 텍스트를 덮어쓰지 않음
"""
from __future__ import annotations
import datetime
import re
from typing import Dict, List, Optional
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

from .attendance_parser import normalize_name, time_to_hours
from .holidays import merge_holidays

# 직원 블록의 행 라벨 (F열에서 검색)
BASIC_LABELS = {"기본", "기본근무", "정상"}

# 명시적 day → col 매핑 (Python/openpyxl 1-based)
# day=1 → col=8 (H), day=2 → col=9 (I), ..., day=31 → col=38 (AL)
DAY_TO_COL: Dict[int, int] = {day: 7 + day for day in range(1, 32)}

# 31일 옆 월말 주휴보정칸
MONTHLY_ADJUSTMENT_COL: int = 39  # AM


def is_formula_cell(cell) -> bool:
    """수식 셀 판별"""
    v = cell.value
    return isinstance(v, str) and v.startswith("=")


def safe_set_value_with_protection(
    ws,
    row: int,
    col: int,
    value,
    log: list,
    review: list,
    employee_name: str,
    date_str: str,
    conflict_kind: str,
) -> bool:
    """
    주휴 등 텍스트 입력 시 사용. 기존 값과 충돌하면 덮어쓰지 않고 검토리스트에 남김.

    빈 셀 또는 같은 값이면 입력. 다른 값이 있으면 검토리스트만.
    수식/병합셀(좌상단 X)은 safe_set_value 와 동일하게 보호됨.
    """
    if not row or row < 1:
        return False
    cell = ws.cell(row=row, column=col)
    coord = f"{get_column_letter(col)}{row}"

    # 병합셀 처리
    if isinstance(cell, MergedCell):
        target = None
        for merged in ws.merged_cells.ranges:
            if cell.coordinate in merged:
                if cell.coordinate != merged.start_cell.coordinate:
                    log.append({
                        "kind": "병합셀_비좌상단",
                        "sheet": ws.title,
                        "cell": coord,
                        "merge_range": str(merged),
                        "skipped_value": value,
                    })
                    return False
                target = ws.cell(row=merged.min_row, column=merged.min_col)
                break
        if target is None:
            return False
        cell = target

    # 수식 셀 보호
    if is_formula_cell(cell):
        log.append({
            "kind": "수식셀_보호",
            "sheet": ws.title,
            "cell": coord,
            "formula": cell.value,
            "skipped_value": value,
        })
        return False

    existing = cell.value
    if existing in (None, "") or existing == value:
        cell.value = value
        return True

    # 충돌 — 덮어쓰지 않고 검토리스트만
    review.append({
        "구분": conflict_kind,
        "성명": employee_name,
        "일자": date_str,
        "요일": "일",
        "PDF원문": "",
        "입력값": f"기존 '{existing}' (덮어쓰지 않음)",
        "메시지": f"{conflict_kind} — {coord} 기존값 보존, 주휴 미입력",
    })
    return False


def safe_set_value(ws, row: int, col: int, value, log: list = None) -> bool:
    """
    안전한 셀 입력. 성공 시 True.
    """
    if not row or row < 1:
        return False
    cell = ws.cell(row=row, column=col)
    coord = f"{get_column_letter(col)}{row}"

    # 병합셀 검사
    if isinstance(cell, MergedCell):
        for merged in ws.merged_cells.ranges:
            if cell.coordinate in merged:
                if cell.coordinate != merged.start_cell.coordinate:
                    if log is not None:
                        log.append({
                            "kind": "병합셀_비좌상단",
                            "sheet": ws.title,
                            "cell": coord,
                            "merge_range": str(merged),
                            "skipped_value": value,
                        })
                    return False
                cell = ws.cell(row=merged.min_row, column=merged.min_col)
                break

    # 수식 셀 보호
    if is_formula_cell(cell):
        if log is not None:
            log.append({
                "kind": "수식셀_보호",
                "sheet": ws.title,
                "cell": coord,
                "formula": cell.value,
                "skipped_value": value,
            })
        return False

    if cell.value is not None and cell.value != "" and log is not None:
        log.append({
            "kind": "기존값_덮어쓰기",
            "sheet": ws.title,
            "cell": coord,
            "old_value": cell.value,
            "new_value": value,
        })

    cell.value = value
    return True


def find_target_sheet(wb, month: int, user_sheet_name: Optional[str]) -> str:
    """대상 근태 시트 자동 감지"""
    sheets = wb.sheetnames

    if user_sheet_name and user_sheet_name in sheets:
        return user_sheet_name

    patterns = [
        rf"근태\s*\(\s*{month}\s*월\s*\)",
        rf"근태\s*{month}\s*월",
        rf"{month}\s*월\s*근태",
    ]
    for ws_name in sheets:
        for p in patterns:
            if re.search(p, ws_name):
                return ws_name

    for ws_name in sheets:
        if "근태" in ws_name:
            return ws_name

    raise ValueError(f"대상 근태 시트를 찾을 수 없습니다 (대상 월: {month}). 사용 가능: {sheets}")


def extract_employee_names_from_sheet(wb, sheet_name: str) -> List[str]:
    """근태 시트에서 직원명 목록 추출 (D열 + F열 '기본' 라벨 기준)"""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    names: List[str] = []
    for r in range(1, ws.max_row + 1):
        d_val = ws.cell(row=r, column=4).value
        f_val = ws.cell(row=r, column=6).value
        if d_val and isinstance(d_val, str) and d_val.strip():
            label = str(f_val).strip() if f_val else ""
            if label in BASIC_LABELS:
                clean = normalize_name(d_val)
                if clean and clean not in names:
                    names.append(clean)
    return names


def find_employee_block(ws, employee_name: str) -> Optional[Dict[str, int]]:
    """
    근태 시트에서 직원 블록 위치를 찾고, 6행 구조로 매핑.

    이 양식에서는 직원당 6행 구조이므로 라벨 검색 대신 위치 강제 매핑:
      basic_row + 0 = 기본
      basic_row + 1 = 연장   ← 텍스트(연차/반차/반반차/유급/주휴) 입력 위치
      basic_row + 2 = 심야
      basic_row + 3 = 특근
      basic_row + 4 = 특잔
      basic_row + 5 = 지각 조퇴
    """
    name_norm = normalize_name(employee_name)
    for r in range(1, ws.max_row + 1):
        d_val = ws.cell(row=r, column=4).value
        f_val = ws.cell(row=r, column=6).value
        if not (d_val and normalize_name(str(d_val)) == name_norm):
            continue
        if not (isinstance(f_val, str) and f_val.strip() in BASIC_LABELS):
            continue

        return {
            "start_row": r,
            "기본": r,
            "연장": r + 1,
            "심야": r + 2,
            "특근": r + 3,
            "특잔": r + 4,
            "지각조퇴": r + 5,
        }
    return None


def get_dow(year: int, month: int, day: int) -> str:
    try:
        d = datetime.date(year, month, day)
        return ["월", "화", "수", "목", "금", "토", "일"][d.weekday()]
    except ValueError:
        return ""


def is_valid_day(year: int, month: int, day: int) -> bool:
    try:
        datetime.date(year, month, day)
        return True
    except ValueError:
        return False


def classify_attendance(slot: dict, year: int, month: int, day: int,
                        paid_holidays: dict, normal_start: str = "08:30",
                        normal_end: str = "17:30") -> dict:
    """
    PDF 슬롯 → 카테고리별 시간 + 비고 + 인정시간

    Returns:
      {
        "기본": 숫자 또는 0,
        "연장": 숫자 또는 0,
        "심야": 숫자 또는 0,
        "특근": 0 (이번 양식에서는 항상 0 — 주말 근무도 연장으로 처리),
        "특잔": 숫자 또는 0,
        "지각조퇴": 음수 또는 0,
        "비고": "" 또는 "연차"/"반차"/"반반차"/"유급",
        "인정시간": 209h 합계 인정 시간 (기본 + 휴가 + 유급 시 8, 그 외 0)
      }
    """
    result = {
        "기본": 0, "연장": 0, "심야": 0, "특근": 0, "특잔": 0,
        "지각조퇴": 0, "비고": "", "인정시간": 0,
    }
    start = slot.get("start", "")
    end = slot.get("end", "")
    ot_str = slot.get("ot", "")
    note = slot.get("note", "")

    if not is_valid_day(year, month, day):
        return result

    if note in ("퇴사",):
        return result

    dow = get_dow(year, month, day)
    is_saturday = dow == "토"
    is_sunday = dow == "일"
    date_str = f"{year}-{month:02d}-{day:02d}"

    try:
        ot_hours = float(ot_str) if ot_str else 0
    except (ValueError, TypeError):
        ot_hours = 0

    # 1. 평일 공휴일 (주말 공휴일은 빈칸)
    if date_str in paid_holidays and not is_saturday and not is_sunday:
        result["기본"] = 8
        result["비고"] = "유급"
        result["인정시간"] = 8
        return result

    # 2. 연차/반차/반반차 → 기본 8 + 비고
    if note in ("연차", "반차", "반반차"):
        result["기본"] = 8
        result["비고"] = note
        result["인정시간"] = 8
        return result

    # 3. 토요일 근무 → 연장 행에 숫자
    if is_saturday and (start or end or ot_hours > 0):
        if ot_hours > 0:
            result["연장"] = ot_hours
        elif start and end:
            normal_dur = max(0, time_to_hours(end) - time_to_hours(start) - 1)
            result["연장"] = normal_dur
        return result

    # 4. 일요일 PDF 근무 → 연장 행 숫자 (주휴는 후처리에서 충돌 방지)
    if is_sunday and (start or end or ot_hours > 0):
        if ot_hours > 0:
            result["연장"] = ot_hours
        elif start and end:
            normal_dur = max(0, time_to_hours(end) - time_to_hours(start) - 1)
            result["연장"] = normal_dur
        return result

    if is_sunday:
        return result

    # 5. 평일 데이터 없음 (결근/빈칸)
    if not start and not end and not note:
        return result

    # 6. 평일 정상 출근
    if start and end:
        s_h = time_to_hours(start)
        e_h = time_to_hours(end)
        normal_s = time_to_hours(normal_start)
        normal_e = time_to_hours(normal_end)

        result["기본"] = 8
        result["인정시간"] = 8

        # 지각/조퇴
        late_hours = max(0, s_h - normal_s)
        early_leave = max(0, normal_e - e_h) if e_h <= normal_e else 0
        if late_hours + early_leave > 0:
            result["지각조퇴"] = -(late_hours + early_leave)

        # 평일 연장
        if e_h > normal_e:
            overtime = ot_hours if ot_hours > 0 else (e_h - normal_e)
            if overtime > 0:
                result["연장"] = overtime

        return result

    # 출근만 있고 퇴근 없음
    if start and not end:
        result["기본"] = 8
        result["인정시간"] = 8
        return result

    return result


def calc_weekly_paid_holiday(
    employee_classified: Dict[int, dict],
    year: int,
    month: int,
    is_employed_in_month: bool = False,
) -> Dict[int, dict]:
    """
    일요일 주휴 자동 판단.

    Args:
      is_employed_in_month: 엑셀 근태표에 이 직원이 month 근무자로 등록되어 있는지

    조건:
      - 일반 일요일: 그 주 월~금 모두 인정 + 다음 주 월요일 인정
      - 월초 첫 일요일 (예: 3/1): 직전 월~금이 이전 달에 있어 검사 불가하므로
        직원이 month 근무자이고 첫째 주(첫 일요일~+7일) 안에 인정 항목이 있으면 인정
      - 월말 마지막 일요일: 다음주 월요일이 다음 달이면 보수적으로 인정

    인정 항목: 정상출근(기본>=8) / 연차 / 반차 / 반반차 / 유급

    Returns: { sunday_day: { "hours": 8.0, "reason": str, "auto": True } }
    """
    sundays_to_pay: Dict[int, dict] = {}

    def has_credit(day: int) -> bool:
        if day < 1 or day > 31:
            return False
        if not is_valid_day(year, month, day):
            return False
        c = employee_classified.get(day, {})
        if c.get("기본", 0) >= 8:
            return True
        if c.get("비고") in ("연차", "반차", "반반차", "유급"):
            return True
        return False

    # 월의 모든 일요일 수집
    sundays_in_month: List[int] = []
    for day in range(1, 32):
        if is_valid_day(year, month, day) and datetime.date(year, month, day).weekday() == 6:
            sundays_in_month.append(day)

    if not sundays_in_month:
        return sundays_to_pay

    first_sunday = sundays_in_month[0]

    for sunday in sundays_in_month:
        # 그 주 월~금 (일요일 기준 2~6일 전 = 금,목,수,화,월)
        # 일요일-1 = 토요일이므로 제외해야 함
        weekdays = [sunday - offset for offset in range(2, 7)]
        weekdays_in_month = [d for d in weekdays if d >= 1 and is_valid_day(year, month, d)]
        weekdays_outside_month_count = 5 - len(weekdays_in_month)

        # 다음 주 월요일
        next_monday = sunday + 1
        next_monday_in_month = is_valid_day(year, month, next_monday)
        next_ok = (not next_monday_in_month) or has_credit(next_monday)

        # 케이스 A: 월초 첫 일요일 (직전 평일이 대부분 이전 달)
        if sunday == first_sunday and weekdays_outside_month_count >= 4:
            # 첫째 주(첫 일요일 포함 + 7일 안)에 인정 항목 있는지
            first_week_end = min(sunday + 7, 31)
            first_week_credit = any(
                has_credit(d) for d in range(1, first_week_end + 1)
                if is_valid_day(year, month, d)
            )
            if is_employed_in_month and first_week_credit:
                sundays_to_pay[sunday] = {
                    "hours": 8.0,
                    "reason": "월초 첫 일요일 — 직원 month 근무자 + 첫째 주 인정 활동 → 주휴",
                    "auto": True,
                }
            continue

        # 케이스 B: 일반 일요일 — 그 주 month 안 평일 모두 인정 + 다음 주 월요일 인정
        weekdays_credit_count = sum(1 for d in weekdays_in_month if has_credit(d))
        if (
            weekdays_in_month
            and weekdays_credit_count == len(weekdays_in_month)
            and next_ok
        ):
            sundays_to_pay[sunday] = {
                "hours": 8.0,
                "reason": (
                    f"월~금 {weekdays_credit_count}일 모두 인정 + "
                    f"{'다음주 월요일 인정' if next_monday_in_month else '다음달 월요일 (보수적 인정)'} → 주휴"
                ),
                "auto": True,
            }

    return sundays_to_pay


def calc_monthly_adjustment(employee_classified: Dict[int, dict],
                             sundays_paid,
                             standard_hours: int = 209) -> float:
    """
    월말 주휴 보정 = 209 - 월합산_인정시간.

    인정시간 = 평일 기본근무(8) + 연차/반차/반반차/유급(8) + 일요일 주휴(8)
    제외 = 평일 연장 + 주말 연장 + 심야 + 특근 + 특잔 + 지각조퇴

    sundays_paid: dict[day, float] 또는 dict[day, {"hours": float, ...}]
    """
    total = 0.0
    for day, c in employee_classified.items():
        rec = c.get("인정시간", 0)
        total += rec
    for v in sundays_paid.values():
        if isinstance(v, dict):
            total += v.get("hours", 0)
        else:
            total += v
    return standard_hours - total


def fill_attendance_sheet(
    excel_path: str,
    pdf_data: Dict[str, Dict[int, dict]],
    year: int,
    month: int,
    sheet_name: Optional[str] = None,
    paid_holidays: Optional[dict] = None,
    standard_hours: int = 209,
    normal_start: str = "08:30",
    normal_end: str = "17:30",
    overwrite_existing: bool = False,
) -> dict:
    """근태 시트 자동 입력 — 6행 구조, 텍스트는 모두 연장 행에 입력"""
    user_holidays = paid_holidays or {}
    merged_holidays = merge_holidays(year, month, user_holidays)

    wb = load_workbook(excel_path, data_only=False)
    sheet_used = find_target_sheet(wb, month, sheet_name)
    ws = wb[sheet_used]

    log: List[dict] = []
    review: List[dict] = []
    stats: Dict[str, int] = {}
    missing: List[str] = []

    counters: Dict[str, int] = {
        "직원_매칭실패": 0,
        "수식셀_덮어쓰기": 0,
        "주말근무_연장행입력": 0,
        "공휴일_유급_입력": 0,
        "주휴_자동입력": 0,
        "연장_텍스트_충돌": 0,
        "일요일주휴_제외": 0,
        "주휴_기본칸_충돌": 0,
        "주휴_표기칸_충돌": 0,
        "유급수정_잘못된주휴": 0,
    }

    for name, days in pdf_data.items():
        block = find_employee_block(ws, name)
        if block is None:
            missing.append(name)
            counters["직원_매칭실패"] += 1
            review.append({
                "구분": "직원_매칭실패",
                "성명": name,
                "메시지": f"엑셀 '{sheet_used}' 시트에서 '{name}' 행을 찾지 못함",
            })
            continue

        classified: Dict[int, dict] = {}
        for day in range(1, 32):
            slot = days.get(day, {})
            classified[day] = classify_attendance(
                slot, year, month, day, merged_holidays, normal_start, normal_end
            )

        # 직원이 매칭됐으니 month 근무자로 등록된 것 → 3/1 같은 월초 주휴 판단 가능
        sundays = calc_weekly_paid_holiday(
            classified, year, month, is_employed_in_month=True
        )

        # 일요일에 PDF 실제 근무가 있으면 주휴 자동 입력 제외
        sundays_filtered: Dict[int, dict] = {}
        for sday, info in sundays.items():
            sunday_c = classified.get(sday, {})
            if sunday_c.get("연장", 0) > 0:
                counters["일요일주휴_제외"] += 1
                review.append({
                    "구분": "주휴_제외",
                    "성명": name,
                    "일자": f"{year}-{month:02d}-{sday:02d}",
                    "요일": "일",
                    "PDF원문": (
                        f"{days.get(sday, {}).get('start', '')} "
                        f"{days.get(sday, {}).get('end', '')} "
                        f"{days.get(sday, {}).get('ot', '')}"
                    ).strip(),
                    "입력값": "(자동 주휴 미입력)",
                    "메시지": "일요일 PDF 실제 근무 발견 → 주휴 자동 입력 제외",
                })
            else:
                sundays_filtered[sday] = info

        # 주휴를 인정시간 합계에 반영하기 위해 sundays_filtered 사용
        cell_count = 0

        # 일자별 셀 입력
        for day in range(1, 32):
            if not is_valid_day(year, month, day):
                continue
            col = DAY_TO_COL[day]  # day=1→8(H), day=2→9(I), ..., day=31→38(AL)
            c = classified[day]
            dow = get_dow(year, month, day)
            is_weekend = dow in ("토", "일")

            # 1. 기본근무 (평일/공휴일/연차)
            if c["기본"]:
                if safe_set_value(ws, block["기본"], col, c["기본"], log):
                    cell_count += 1

            # 2. 연장 행 — 시간(숫자) 또는 텍스트 — 동시 발생 시 숫자 우선
            ot_value = c["연장"]
            note_text = c.get("비고")

            if ot_value and note_text:
                # 충돌: 평일 연장 + 연차/반차/반반차/유급 동시 발생
                counters["연장_텍스트_충돌"] += 1
                review.append({
                    "구분": "충돌_연장숫자vs텍스트",
                    "성명": name,
                    "일자": f"{year}-{month:02d}-{day:02d}",
                    "요일": dow,
                    "PDF원문": (
                        f"{days.get(day, {}).get('start', '')} "
                        f"{days.get(day, {}).get('end', '')} "
                        f"{days.get(day, {}).get('ot', '')} "
                        f"{days.get(day, {}).get('note', '')}"
                    ).strip(),
                    "입력값": f"연장 {ot_value}h",
                    "메시지": f"평일 연장 {ot_value}h vs 비고 '{note_text}' 충돌 — 숫자 우선",
                })
                if safe_set_value(ws, block["연장"], col, ot_value, log):
                    cell_count += 1
            elif ot_value:
                # 연장 숫자만 (평일 연장 또는 주말 근무)
                if safe_set_value(ws, block["연장"], col, ot_value, log):
                    cell_count += 1
                if is_weekend:
                    counters["주말근무_연장행입력"] += 1
                    review.append({
                        "구분": "주말근무_연장행",
                        "성명": name,
                        "일자": f"{year}-{month:02d}-{day:02d}",
                        "요일": dow,
                        "PDF원문": (
                            f"{days.get(day, {}).get('start', '')} "
                            f"{days.get(day, {}).get('end', '')} "
                            f"{days.get(day, {}).get('ot', '')}"
                        ).strip(),
                        "입력값": f"연장 {ot_value}h",
                        "메시지": "주말 근무 → 연장 행 (특근 행 X), 209h 인정 제외",
                    })
            elif note_text:
                # 비고 텍스트만 (연차/반차/반반차/유급)
                if safe_set_value(ws, block["연장"], col, note_text, log):
                    cell_count += 1
                if note_text == "유급":
                    counters["공휴일_유급_입력"] += 1
                    review.append({
                        "구분": "공휴일_유급",
                        "성명": name,
                        "일자": f"{year}-{month:02d}-{day:02d}",
                        "요일": dow,
                        "PDF원문": days.get(day, {}).get("note", ""),
                        "입력값": "기본 8 + 연장행 '유급'",
                        "메시지": f"{merged_holidays.get(f'{year}-{month:02d}-{day:02d}', '공휴일')} 자동 유급 처리",
                    })
                else:
                    review.append({
                        "구분": f"비고_{note_text}",
                        "성명": name,
                        "일자": f"{year}-{month:02d}-{day:02d}",
                        "요일": dow,
                        "PDF원문": days.get(day, {}).get("note", ""),
                        "입력값": f"기본 8 + 연장행 '{note_text}'",
                        "메시지": f"{note_text} 처리 — 기본 8h 인정 + 연장 행에 텍스트 표기",
                    })

            # 3. 심야
            if c["심야"]:
                if safe_set_value(ws, block["심야"], col, c["심야"], log):
                    cell_count += 1

            # 4. 특잔
            if c["특잔"]:
                if safe_set_value(ws, block["특잔"], col, c["특잔"], log):
                    cell_count += 1

            # 5. 지각/조퇴
            if c["지각조퇴"]:
                if safe_set_value(ws, block["지각조퇴"], col, c["지각조퇴"], log):
                    cell_count += 1

        # 일요일 주휴 입력 — 기본 행에 8 + 연장 행에 "주휴"
        # 안전장치:
        #   1. day가 진짜 일요일인지 재검증 (평일/토요일에 주휴 입력 방지)
        #   2. 평일 공휴일(예: 3/2 삼일절 대체)에는 절대 주휴 입력 금지
        #   3. 기존 값과 충돌 시 덮어쓰지 않고 검토리스트에 남김
        for sday, info in sundays_filtered.items():
            # 안전장치 1: day가 일요일인지 재검증
            if not is_valid_day(year, month, sday):
                continue
            sd = datetime.date(year, month, sday)
            if sd.weekday() != 6:
                review.append({
                    "구분": "주휴_방어_일요일아님",
                    "성명": name,
                    "일자": f"{year}-{month:02d}-{sday:02d}",
                    "요일": "월화수목금토일"[sd.weekday()],
                    "PDF원문": "",
                    "입력값": "(주휴 입력 거부)",
                    "메시지": f"day={sday}는 일요일이 아님 → 주휴 입력 거부",
                })
                continue

            # 안전장치 2: 평일 공휴일 날짜에는 주휴 입력 금지
            sday_date_str = f"{year}-{month:02d}-{sday:02d}"
            if sday_date_str in merged_holidays:
                # 일요일 + 공휴일 동시 — 주휴를 우선 (3/1 삼일절+일요일 같은 케이스)
                # 사용자 요구: 일요일 공휴일은 빈칸 또는 주휴, 평일 공휴일은 유급
                # 이미 sd.weekday()==6 보장됨 → 그대로 진행
                pass

            scol = DAY_TO_COL[sday]  # day=1→8(H), day=8→15(O), day=15→22(V), day=22→29(AC), day=29→36(AJ)
            hrs = info["hours"]
            reason = info["reason"]
            date_str = sday_date_str

            # 기본 칸 (8) — 보호 모드
            basic_ok = safe_set_value_with_protection(
                ws, block["기본"], scol, hrs, log, review,
                employee_name=name, date_str=date_str,
                conflict_kind="주휴_기본칸_충돌",
            )
            if basic_ok:
                cell_count += 1
            else:
                counters["주휴_기본칸_충돌"] += 1

            # 연장 칸 ("주휴") — 보호 모드
            note_ok = safe_set_value_with_protection(
                ws, block["연장"], scol, "주휴", log, review,
                employee_name=name, date_str=date_str,
                conflict_kind="주휴_표기칸_충돌",
            )
            if note_ok:
                cell_count += 1
            else:
                counters["주휴_표기칸_충돌"] += 1

            counters["주휴_자동입력"] += 1
            review.append({
                "구분": "주휴_자동입력",
                "성명": name,
                "일자": date_str,
                "요일": "일",
                "PDF원문": "",
                "입력값": f"기본 {hrs} + 연장행 '주휴'",
                "메시지": reason,
            })

        # === 평일 공휴일 자가 교정 sweep ===
        # 안전장치: 평일 공휴일 칸의 연장 행에 잘못된 "주휴"가 있으면 "유급"으로 교체.
        # (이전 처리 또는 어떤 이유로 잘못 입력된 경우 방어)
        for date_str_h in merged_holidays:
            parts = date_str_h.split("-")
            if len(parts) != 3:
                continue
            try:
                hday = int(parts[2])
                hd = datetime.date(year, month, hday)
            except (ValueError, IndexError):
                continue
            # 주말 공휴일은 빈칸이므로 sweep 대상 X
            if hd.weekday() in (5, 6):
                continue
            hcol = DAY_TO_COL.get(hday)
            if not hcol:
                continue

            note_cell = ws.cell(row=block["연장"], column=hcol)
            # 병합셀/수식셀 보호
            if isinstance(note_cell, MergedCell):
                continue
            if is_formula_cell(note_cell):
                continue

            existing = note_cell.value
            if existing == "주휴":
                note_cell.value = "유급"
                counters["유급수정_잘못된주휴"] += 1
                cell_count += 1
                review.append({
                    "구분": "유급수정",
                    "성명": name,
                    "일자": date_str_h,
                    "요일": "월화수목금"[hd.weekday()],
                    "PDF원문": "",
                    "입력값": "기존 '주휴' → '유급'",
                    "메시지": (
                        f"{merged_holidays[date_str_h]} 평일 공휴일 — "
                        f"잘못 입력된 '주휴'를 '유급'으로 자가 수정"
                    ),
                })

        # 월말 주휴 보정 — 31일 옆 (AM=39)
        adjustment = calc_monthly_adjustment(classified, sundays_filtered, standard_hours)
        if abs(adjustment) > 0.01:
            adj_col = MONTHLY_ADJUSTMENT_COL  # 39 (AM)
            adj_value = round(adjustment, 1)
            # 보정값은 기본 행에 입력 (양식에 별도 주휴 행이 없으므로)
            if safe_set_value(ws, block["기본"], adj_col, adj_value, log):
                cell_count += 1
            인정합계 = standard_hours - adj_value
            review.append({
                "구분": "월말_보정",
                "성명": name,
                "일자": f"{year}-{month:02d}-말",
                "요일": "",
                "PDF원문": "",
                "입력값": str(adj_value),
                "메시지": f"인정시간 {인정합계}h → 보정 {adj_value}h (= 209 - 인정합계)",
            })

        stats[name] = cell_count

    # 수식셀 보호 카운터
    for entry in log:
        if entry.get("kind") == "수식셀_보호":
            counters["수식셀_덮어쓰기"] += 1

    return {
        "stats": stats,
        "missing": missing,
        "log": log,
        "review_list": review,
        "wb": wb,
        "sheet_used": sheet_used,
        "year": year,
        "month": month,
        "merged_holidays": merged_holidays,
        "counters": counters,
    }
