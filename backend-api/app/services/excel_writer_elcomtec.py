"""
엘컴텍 근태 시트 자동 입력 + 연차/반차 자동판단 (v2 — 출퇴근시간 기준)

지각/조퇴는 PDF 인정근무시간이 아니라 **출근/퇴근 시각** 기준으로 계산:
  기준 08:30~17:30, 기본 8h, 점심 12:30~13:30.
  실근무 = (min(퇴근,17:30) - max(출근,08:30)) - (그 구간과 겹치는 점심)
  부족 = 8 - 실근무 → 30분 단위 올림 → 지각조퇴 = -부족
  · 늦게까지 일해도 지각 상쇄 안 함(퇴근은 17:30로 cap, 초과분은 연장행)
연장/심야 = PDF 평일연장/야간 그대로.

연차 자동(월 최대 1개): 전체월 재직 & 중도입퇴사 아님 & 잔여>0 & 평일결근 & 당월 연차<1
  → 기본8/연장=연차, 연차내역 4월 +1. 추가 결근은 결근(빈칸).
반차 자동(월 최대 2건): 전체월 재직 & 중도입퇴사 아님 & 잔여≥0.5 & 조퇴 부족 4~4.5h
  → 기본8/연장=반차/지각조퇴 빈칸, 연차내역 4월 +0.5.
중도입사(hire>월초)/중도퇴사(resign<월말 또는 PDF기록이 월말 전 종료) → 연차·반차 자동 제외.
주휴: 연차 처리 결근=인정(포함) / 비연차 결근=제외.
검토시트 '연차_반차_후보검토' 생성. 수식·서식 100% 보존.
"""
from __future__ import annotations
import datetime
import math
import calendar
from typing import Optional

from openpyxl import load_workbook

from .excel_writer import (
    DAY_TO_COL,
    find_target_sheet,
    build_name_to_row_map,
    find_employee_block,
    safe_set_value,
    safe_set_value_with_protection,
    get_dow,
    normalize_name,
)

LEAVE_SHEET = "연차내역"
DB_SHEET = "DB"
REVIEW_SHEET = "연차_반차_후보검토"
MAX_ANNUAL_PER_MONTH = 1
MAX_BANCHA_PER_MONTH = 2
STD_START, STD_END = 8.5, 17.5      # 08:30 ~ 17:30
LUNCH_START, LUNCH_END = 12.5, 13.5  # 12:30 ~ 13:30
BANCHA_MIN, BANCHA_MAX = 4.0, 4.5    # 조퇴 반차 인정 부족시간 범위


def _ceil_half(h: float) -> float:
    return math.ceil(round(h, 4) * 2) / 2


def _overlap(s: float, e: float, ls: float, le: float) -> float:
    return max(0.0, min(e, le) - max(s, ls))


def _shortfall(start: Optional[float], end: Optional[float]) -> float:
    """출/퇴근 시각(소수시간) → 부족시간(30분 올림). 늦은 퇴근은 17:30로 cap."""
    if start is None or end is None:
        return 0.0
    s = max(start, STD_START)
    e = min(end, STD_END)
    worked = max(0.0, e - s) - _overlap(s, e, LUNCH_START, LUNCH_END)
    sf = 8.0 - worked
    return _ceil_half(sf) if sf > 0.0001 else 0.0


def _to_date(v) -> Optional[datetime.date]:
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    return None


def _read_db_dates(path: str) -> dict:
    out: dict = {}
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return out
    try:
        if DB_SHEET not in wb.sheetnames:
            return out
        ws = wb[DB_SHEET]
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[1] if len(row) > 1 else None
            if not name or not isinstance(name, str):
                continue
            out[normalize_name(name)] = {
                "hire": _to_date(row[4]) if len(row) > 4 else None,
                "resign": _to_date(row[5]) if len(row) > 5 else None,
            }
    finally:
        wb.close()
    return out


def _read_leave_ledger(path: str) -> dict:
    out: dict = {}
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return out
    try:
        if LEAVE_SHEET not in wb.sheetnames:
            return out
        ws = wb[LEAVE_SHEET]
        T, L, C = 20, 12, 3
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=5).value != "발생":
                continue
            name = ws.cell(row=r, column=C).value
            if not name or not isinstance(name, str):
                continue

            def num(v):
                try:
                    return float(v) if v is not None else 0.0
                except Exception:
                    return 0.0
            발생 = num(ws.cell(row=r, column=T).value)
            사용 = num(ws.cell(row=r + 1, column=T).value)
            april_cur = num(ws.cell(row=r + 1, column=L).value)
            out[normalize_name(name)] = {
                "used_row": r + 1,
                "april_col": L,
                "available": 발생 - (사용 - april_cur),  # 4월 사용 가능 총량
            }
    finally:
        wb.close()
    return out


def _is_unexcused_absent(rec: Optional[dict], day: int, annual_days: set) -> bool:
    if day in annual_days:
        return False  # 연차 = 근태 인정
    if rec is None:
        return True
    if rec.get("구분") == "평일" and not rec.get("평_기본"):
        return True
    return False


def _week_full_attendance(days: dict, sunday: datetime.date, month: int, annual_days: set) -> bool:
    monday = sunday - datetime.timedelta(days=6)
    for off in range(5):  # Mon..Fri
        wd = monday + datetime.timedelta(days=off)
        if wd.month != month:
            continue
        if _is_unexcused_absent(days.get(wd.day), wd.day, annual_days):
            return False
    return True


def fill_attendance_sheet_elcomtec(
    excel_path: str,
    parsed: dict,
    year: int,
    month: int,
    sheet_name: Optional[str] = None,
    name_to_row: Optional[dict] = None,
    leave_ledger: Optional[dict] = None,
    db_dates: Optional[dict] = None,
) -> dict:
    """엘컴텍 근태 시트 자동 입력 + 연차/반차. excel_path in-place 수정, wb 반환."""
    if leave_ledger is None:
        leave_ledger = _read_leave_ledger(excel_path)
    if db_dates is None:
        db_dates = _read_db_dates(excel_path)

    wb = load_workbook(excel_path, data_only=False, keep_links=False)
    sheet_used = find_target_sheet(wb, month, sheet_name)
    ws = wb[sheet_used]
    ws_leave = wb[LEAVE_SHEET] if LEAVE_SHEET in wb.sheetnames else None
    if name_to_row is None:
        name_to_row = build_name_to_row_map(excel_path, sheet_used)

    month_start = datetime.date(year, month, 1)
    last_day = calendar.monthrange(year, month)[1]
    month_end = datetime.date(year, month, last_day)

    log: list = []
    review: list = []
    candidate_rows: list = []
    stats: dict = {}
    missing: list = []
    counters = {
        "직원_매칭실패": 0, "지각조퇴_입력": 0, "연차_자동적용": 0, "반차_자동적용": 0,
        "미인정결근": 0, "중도입퇴사_연차반차제외": 0, "주휴_자동입력": 0,
        "토요일_연장입력": 0, "일요일_특근입력": 0,
    }

    for name, days in parsed.items():
        block = find_employee_block(ws, name, name_to_row)
        if not block:
            counters["직원_매칭실패"] += 1
            missing.append(name)
            review.append({"구분": "매칭실패", "성명": name, "일자": "", "요일": "",
                           "PDF원문": "", "입력값": "", "메시지": f"'{name}' 블록 없음(이름 확인)"})
            continue

        norm = normalize_name(name)
        r_기본 = block["기본"]; r_연장 = block["연장"]; r_심야 = block["심야"]
        r_특근 = block["특근"]; r_특잔 = block["특잔"]; r_지각조퇴 = block["지각조퇴"]

        info = db_dates.get(norm, {})
        hire = info.get("hire"); resign = info.get("resign")
        led = leave_ledger.get(norm, {})
        available = led.get("available", 0.0)

        day_nums = sorted(days)
        pdf_last = max(day_nums) if day_nums else last_day
        is_mid_joiner = bool(hire and hire > month_start)
        is_mid_leaver = bool((resign and resign < month_end) or (pdf_last < last_day))
        full_month = not is_mid_joiner and not is_mid_leaver

        filled = 0
        used_quota = 0.0   # 이번 실행 연차/반차 소모(잔여 차감)
        annual_cnt = 0
        bancha_cnt = 0
        annual_days: set = set()

        for day in day_nums:
            rec = days[day]
            col = DAY_TO_COL.get(day)
            if not col:
                continue
            gb = rec.get("구분"); ot = rec.get("출결")
            date_str = f"{year}-{month:02d}-{day:02d}"
            dow = get_dow(year, month, day)
            day_date = datetime.date(year, month, day)
            start = rec.get("인정출근") if rec.get("인정출근") is not None else rec.get("실출근")
            end = rec.get("인정퇴근") if rec.get("인정퇴근") is not None else rec.get("실퇴근")

            if gb == "평일":
                base = rec.get("평_기본")
                if base:  # 근무
                    if safe_set_value(ws, r_기본, col, 8, log):
                        filled += 1
                    if rec.get("평_연장") and safe_set_value(ws, r_연장, col, rec["평_연장"], log):
                        filled += 1
                    if rec.get("평_야간") and safe_set_value(ws, r_심야, col, rec["평_야간"], log):
                        filled += 1
                    sf = _shortfall(start, end)
                    is_jotwe = (ot == "조퇴") or (end is not None and end <= 13.6)
                    # 반차: 전체월재직 + 잔여≥0.5 + 조퇴 부족 4~4.5h + 월≤2건
                    if (full_month and is_jotwe and BANCHA_MIN <= sf <= BANCHA_MAX
                            and bancha_cnt < MAX_BANCHA_PER_MONTH and (available - used_quota) >= 0.5):
                        safe_set_value_with_protection(ws, r_연장, col, "반차", log, review, name, date_str, "반차_충돌")
                        bancha_cnt += 1; used_quota += 0.5; filled += 1
                        counters["반차_자동적용"] += 1
                        candidate_rows.append({
                            "직원명": name, "날짜": date_str, "PDF출결": ot or "조퇴",
                            "인정근무시간": rec.get("간") or "", "지각시간": rec.get("지각") or "", "조퇴시간": rec.get("조퇴") or "",
                            "평일주말": "평일", "입사일": str(hire or ""), "퇴사일": str(resign or ""),
                            "연차잔여": round(available - used_quota, 2), "자동판단": "반차",
                            "입력값": "기본8/연장=반차/지각조퇴 빈칸", "검토메시지": f"조퇴 부족 {sf}h(반차 수준) → 반차 자동, 연차내역 -0.5"})
                    elif sf > 0:
                        if safe_set_value(ws, r_지각조퇴, col, -sf, log):
                            filled += 1; counters["지각조퇴_입력"] += 1
                        candidate_rows.append({
                            "직원명": name, "날짜": date_str, "PDF출결": ot or "지각/조퇴",
                            "인정근무시간": rec.get("간") or "", "지각시간": rec.get("지각") or "", "조퇴시간": rec.get("조퇴") or "",
                            "평일주말": "평일", "입사일": str(hire or ""), "퇴사일": str(resign or ""),
                            "연차잔여": round(available - used_quota, 2), "자동판단": "지각조퇴",
                            "입력값": f"지각조퇴={-sf}", "검토메시지": f"출근{rec.get('인정출근')}/퇴근{rec.get('인정퇴근')} 기준 부족 {sf}h"})
                else:  # 평일 결근
                    if is_mid_joiner or is_mid_leaver:
                        counters["중도입퇴사_연차반차제외"] += 1
                        candidate_rows.append({
                            "직원명": name, "날짜": date_str, "PDF출결": ot or "결근",
                            "인정근무시간": "", "지각시간": "", "조퇴시간": "",
                            "평일주말": "평일", "입사일": str(hire or ""), "퇴사일": str(resign or ""),
                            "연차잔여": round(available - used_quota, 2), "자동판단": "중도입퇴사-제외",
                            "입력값": "(미입력)", "검토메시지": "중도입사/퇴사자 → 연차·반차 자동처리 제외(결근/빈칸)"})
                    elif full_month and annual_cnt < MAX_ANNUAL_PER_MONTH and (available - used_quota) >= 1:
                        if safe_set_value(ws, r_기본, col, 8, log):
                            filled += 1
                        if safe_set_value_with_protection(ws, r_연장, col, "연차", log, review, name, date_str, "연차_충돌"):
                            filled += 1
                        annual_cnt += 1; used_quota += 1; annual_days.add(day)
                        counters["연차_자동적용"] += 1
                        candidate_rows.append({
                            "직원명": name, "날짜": date_str, "PDF출결": ot or "결근",
                            "인정근무시간": "", "지각시간": "", "조퇴시간": "",
                            "평일주말": "평일", "입사일": str(hire or ""), "퇴사일": str(resign or ""),
                            "연차잔여": round(available - used_quota, 2), "자동판단": "연차",
                            "입력값": "기본8/연장=연차", "검토메시지": "평일결근+잔여연차 → 연차 자동(월 1개 한도), 연차내역 -1"})
                    else:
                        counters["미인정결근"] += 1
                        why = "월 연차한도(1) 초과" if annual_cnt >= MAX_ANNUAL_PER_MONTH else ("잔여연차 없음" if (available - used_quota) < 1 else "보류")
                        candidate_rows.append({
                            "직원명": name, "날짜": date_str, "PDF출결": ot or "결근",
                            "인정근무시간": "", "지각시간": "", "조퇴시간": "",
                            "평일주말": "평일", "입사일": str(hire or ""), "퇴사일": str(resign or ""),
                            "연차잔여": round(available - used_quota, 2), "자동판단": "결근",
                            "입력값": "(미입력)", "검토메시지": f"연차 미적용({why}) → 결근(주휴 제외)"})

            elif gb == "무휴":  # 토요일
                hb = rec.get("휴_기본")
                if hb:
                    if safe_set_value(ws, r_연장, col, hb, log):
                        filled += 1; counters["토요일_연장입력"] += 1
                    if rec.get("휴_연장") and safe_set_value(ws, r_특잔, col, rec["휴_연장"], log):
                        filled += 1

            elif gb == "주휴":  # 일요일 근무
                hb = rec.get("휴_기본")
                if hb:
                    if safe_set_value(ws, r_특근, col, hb, log):
                        filled += 1; counters["일요일_특근입력"] += 1
                    if rec.get("휴_연장") and safe_set_value(ws, r_특잔, col, rec["휴_연장"], log):
                        filled += 1

        # 주휴 개근(연차 인정 포함)
        for day in day_nums:
            if days[day].get("구분") != "주휴":
                continue
            sunday = datetime.date(year, month, day)
            if not _week_full_attendance(days, sunday, month, annual_days):
                continue
            col = DAY_TO_COL.get(day)
            if not col:
                continue
            if safe_set_value(ws, r_기본, col, 8, log):
                filled += 1
            if safe_set_value_with_protection(ws, r_연장, col, "주휴", log, review, name, f"{year}-{month:02d}-{day:02d}", "주휴_충돌"):
                filled += 1; counters["주휴_자동입력"] += 1

        # 연차내역 4월 사용란 = 연차*1 + 반차*0.5
        total_use = annual_cnt * 1.0 + bancha_cnt * 0.5
        if total_use > 0 and ws_leave and led.get("used_row"):
            safe_set_value(ws_leave, led["used_row"], led["april_col"], total_use, log)

        stats[name] = filled

    # 검토 시트
    if REVIEW_SHEET in wb.sheetnames:
        del wb[REVIEW_SHEET]
    ws_rev = wb.create_sheet(REVIEW_SHEET)
    headers = ["직원명", "날짜", "PDF출결", "인정근무시간", "지각시간", "조퇴시간",
               "평일/주말", "입사일", "퇴사일", "연차잔여", "자동판단", "입력값", "검토메시지"]
    ws_rev.append(headers)
    for c in candidate_rows:
        ws_rev.append([c["직원명"], c["날짜"], c["PDF출결"], c["인정근무시간"], c["지각시간"], c["조퇴시간"],
                       c["평일주말"], c["입사일"], c["퇴사일"], c["연차잔여"], c["자동판단"], c["입력값"], c["검토메시지"]])

    return {
        "wb": wb, "sheet_used": sheet_used, "stats": stats, "missing": missing,
        "review_list": review, "candidate_rows": candidate_rows, "log": log, "counters": counters,
    }
