"""
로컬 처리 + 완성 폼 비교 — 사용자 환경에서 직접 실행
"""
import sys
import io
import os
import shutil
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

INPUT_DIR = "C:/Users/USER/OneDrive/Desktop/근퇴 완성"
PDF_PATH = f"{INPUT_DIR}/(주)엘티와이 출근부_26.03월_더웰.pdf"
SOURCE_XLSX = f"{INPUT_DIR}/3월청구(엘티와이).xlsx"
GOLDEN_XLSX = f"{INPUT_DIR}/근퇴 테스트.xlsx"
OUT_DIR = "C:/tmp/lty_test"
os.makedirs(OUT_DIR, exist_ok=True)
OUT_XLSX = f"{OUT_DIR}/result.xlsx"

YEAR = 2026
MONTH = 3

print("=" * 70)
print(f"INPUT")
print("=" * 70)
print(f"  PDF:    {PDF_PATH} ({os.path.getsize(PDF_PATH)/1024:.1f} KB)")
print(f"  XLSX:   {SOURCE_XLSX} ({os.path.getsize(SOURCE_XLSX)/1024:.1f} KB)")
print(f"  GOLDEN: {GOLDEN_XLSX} ({os.path.getsize(GOLDEN_XLSX)/1024:.1f} KB)")
print(f"  OUT:    {OUT_XLSX}")
print()

# === Step 1: 원본 파일 복사 ===
print("=" * 70)
print(f"Step 1: Copy source → output")
print("=" * 70)
shutil.copyfile(SOURCE_XLSX, OUT_XLSX)
print(f"OK: {os.path.getsize(OUT_XLSX)/1024:.1f} KB")
print()

# === Step 2: 엑셀 사전 로드 ===
print("=" * 70)
print(f"Step 2: Excel pre-load")
print("=" * 70)
from openpyxl import load_workbook
from app.services.excel_writer import find_target_sheet, extract_employee_names_from_sheet, DAY_TO_COL
from app.services.hire_date import extract_employee_info_from_db, find_feb_attendance_sheet, check_feb_last_week_credit

wb_pre_values = load_workbook(SOURCE_XLSX, data_only=True, read_only=True)
wb_pre = load_workbook(SOURCE_XLSX, data_only=False, read_only=False)
march_sheet = find_target_sheet(wb_pre, MONTH, None)
print(f"  march_sheet: {march_sheet}")

# data_only=True wb에서 직원명 추출 (D열 VLOOKUP 평가)
excel_names = extract_employee_names_from_sheet(wb_pre_values, march_sheet)
print(f"  excel_names: {len(excel_names)}명: {excel_names[:5]}...")

db_info = extract_employee_info_from_db(wb_pre_values)
print(f"  db_info: {len(db_info)}명")

feb_sheet = find_feb_attendance_sheet(wb_pre_values, YEAR, MONTH)
print(f"  feb_sheet: {feb_sheet}")

# 입사일 / 퇴사일 매핑
hire_dates = {}
resign_dates = {}
for norm, info in db_info.items():
    if info.get("hire_date"):
        hire_dates[norm] = info["hire_date"]
    if info.get("resign_date"):
        resign_dates[norm] = info["resign_date"]
print(f"  hire_dates: {len(hire_dates)}명")
print(f"  resign_dates: {len(resign_dates)}명")

# 2월 인정 검사
feb_credit_map = {}
if feb_sheet:
    for nm in excel_names:
        feb_credit_map[nm] = check_feb_last_week_credit(
            wb_pre_values, feb_sheet, nm, YEAR, feb_month=MONTH-1, day_to_col=DAY_TO_COL
        )
wb_pre.close()
wb_pre_values.close()
print()

# === Step 3: PDF 파싱 ===
print("=" * 70)
print(f"Step 3: PDF parsing")
print("=" * 70)
from app.services.attendance_parser import parse_pdf

pdf_data, meta = parse_pdf(PDF_PATH, expected_names=excel_names)
print(f"  pdf_data: {len(pdf_data)}명")
print(f"  matched_employees: {len(set(meta.get('matched_employees', [])))}명")
print(f"  pages: {meta.get('raw_pages')}")
print(f"  day_corrections: {len(meta.get('day_corrections', []))}")
print(f"  Names in pdf_data:")
for name in list(pdf_data.keys())[:15]:
    print(f"    - {name}")
if len(pdf_data) > 15:
    print(f"    ... +{len(pdf_data)-15} more")
print()

# === Step 4: 자동 입력 ===
print("=" * 70)
print(f"Step 4: fill_attendance_sheet")
print("=" * 70)
from app.services.excel_writer import fill_attendance_sheet

# 추정 입사일 (PDF에서 보완)
from app.services.hire_date import estimate_hire_dates_from_pdf, merge_hire_dates
hire_pdf = estimate_hire_dates_from_pdf(pdf_data, YEAR, MONTH)
hire_merged = merge_hire_dates(hire_dates, hire_pdf)
print(f"  hire (DB): {len(hire_dates)} + PDF estimate: {len(hire_pdf)} = merged: {len(hire_merged)}")

att_result = fill_attendance_sheet(
    OUT_XLSX,
    pdf_data,
    year=YEAR,
    month=MONTH,
    paid_holidays={},
    standard_hours=209,
    normal_start="08:30",
    normal_end="17:30",
    hire_dates=hire_merged,
    feb_credit_map=feb_credit_map,
    resign_dates=resign_dates,
)
print(f"  stats: {len(att_result['stats'])}명")
print(f"  missing: {len(att_result['missing'])}명: {att_result['missing'][:5]}...")
print(f"  total cells: {sum(att_result['stats'].values())}")
print()

# Save
wb = att_result["wb"]
wb.save(OUT_XLSX)
print(f"Saved: {OUT_XLSX} ({os.path.getsize(OUT_XLSX)/1024:.1f} KB)")
print()
